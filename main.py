"""
LP Notice Analyzer - Backend Server
Run: uvicorn main:app --reload --port 8000
"""
import os, json, base64, io, time, hashlib, asyncio
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional
import httpx
from supabase import create_client

# ── Config ──────────────────────────────────────────────
GEMINI_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-3-flash-preview"
GEMINI_TEMPERATURE = 0.1
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "")

# ── App ─────────────────────────────────────────────────
app = FastAPI(title="LP Notice Analyzer API")
app.add_middleware(CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# Serve frontend — index.html in same directory as main.py
FRONTEND_DIR = Path(__file__).resolve().parent
_index_html = FRONTEND_DIR / "index.html"
print(f"  [INIT] Frontend path: {_index_html} (exists: {_index_html.exists()})")

@app.get("/")
async def root_redirect():
    return RedirectResponse(url="/app")

@app.get("/app")
async def serve_app():
    if _index_html.exists():
        return FileResponse(str(_index_html), media_type="text/html")
    # Fallback: try current working directory
    cwd_html = Path.cwd() / "index.html"
    if cwd_html.exists():
        return FileResponse(str(cwd_html), media_type="text/html")
    raise HTTPException(404, f"index.html not found. Checked: {_index_html}, {cwd_html}")

@app.get("/api/auth/config")
async def auth_config():
    """Return Supabase public config for frontend auth. No auth required."""
    return {"supabase_url": SUPABASE_URL, "supabase_anon_key": SUPABASE_ANON_KEY}

# ── Database (Supabase) ────────────────────────────────
_supabase = None
def get_supa():
    global _supabase
    if _supabase is None:
        if not SUPABASE_URL or not SUPABASE_KEY:
            raise RuntimeError("SUPABASE_URL and SUPABASE_KEY environment variables must be set")
        _supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        print(f"  [INIT] Supabase connected: {SUPABASE_URL[:40]}...")
    return _supabase

def db_get(table, id_val, id_col="id"):
    r = get_supa().table(table).select("*").eq(id_col, id_val).limit(1).execute()
    return r.data[0] if r.data else None

def db_list(table, order_col=None, order_desc=True, **filters):
    q = get_supa().table(table).select("*")
    for k, v in filters.items():
        q = q.eq(k, v)
    if order_col:
        q = q.order(order_col, desc=order_desc)
    return q.execute().data or []

def db_insert(table, data):
    get_supa().table(table).insert(data).execute()

def db_upsert(table, data):
    get_supa().table(table).upsert(data).execute()

def db_update(table, data, id_val, id_col="id"):
    get_supa().table(table).update(data).eq(id_col, id_val).execute()

def db_delete(table, id_val, id_col="id"):
    get_supa().table(table).delete().eq(id_col, id_val).execute()

def db_find(table, col, val):
    r = get_supa().table(table).select("*").eq(col, val).limit(1).execute()
    return r.data[0] if r.data else None

def db_count(table):
    r = get_supa().table(table).select("id", count="exact").execute()
    return r.count or 0

print(f"  [INIT] DB: Supabase ({SUPABASE_URL[:30]}...)" if SUPABASE_URL else "  [WARN] SUPABASE_URL not set!")

def _get_setting(key, default=None, org_id=None):
    q = get_supa().table("settings").select("*").eq("key", key)
    if org_id:
        q = q.eq("org_id", org_id)
    r = q.limit(1).execute()
    return r.data[0]["value"] if r.data else default

def _set_setting(key, value, org_id=None):
    q = get_supa().table("settings").select("*").eq("key", key)
    if org_id:
        q = q.eq("org_id", org_id)
    existing = q.limit(1).execute()
    if existing.data:
        uq = get_supa().table("settings").update({"value": value}).eq("key", key)
        if org_id:
            uq = uq.eq("org_id", org_id)
        uq.execute()
    else:
        data = {"key": key, "value": value}
        if org_id:
            data["org_id"] = org_id
        get_supa().table("settings").insert(data).execute()


# ── Storage (Supabase Storage) ─────────────────────────
def store_pdf(notice_id: str, pdf_bytes: bytes):
    """Save PDF to Supabase Storage."""
    try:
        get_supa().storage.from_("pdfs").upload(
            f"{notice_id}.pdf", pdf_bytes,
            {"content-type": "application/pdf", "upsert": "true"})
    except Exception as e:
        if "Duplicate" in str(e) or "already exists" in str(e):
            get_supa().storage.from_("pdfs").update(
                f"{notice_id}.pdf", pdf_bytes,
                {"content-type": "application/pdf"})
        else:
            raise

def load_pdf(notice_id: str) -> bytes:
    """Load PDF from Supabase Storage. Returns bytes or None."""
    try:
        return get_supa().storage.from_("pdfs").download(f"{notice_id}.pdf")
    except Exception:
        return None

def pdf_exists(notice_id: str) -> bool:
    """Check if PDF exists in storage."""
    return load_pdf(notice_id) is not None

def store_text_map(notice_id: str, text_map: list):
    """Save text_map.json to Supabase Storage."""
    data = json.dumps(text_map, ensure_ascii=False).encode("utf-8")
    try:
        get_supa().storage.from_("pages").upload(
            f"{notice_id}/text_map.json", data,
            {"content-type": "application/json", "upsert": "true"})
    except Exception as e:
        if "Duplicate" in str(e) or "already exists" in str(e):
            get_supa().storage.from_("pages").update(
                f"{notice_id}/text_map.json", data,
                {"content-type": "application/json"})
        else:
            raise

def load_text_map(notice_id: str) -> list:
    """Load text_map.json from Supabase Storage. Returns list or []."""
    try:
        data = get_supa().storage.from_("pages").download(f"{notice_id}/text_map.json")
        return json.loads(data.decode("utf-8")) if data else []
    except Exception:
        return []

def delete_storage(notice_id: str):
    """Delete PDF + text_map from Supabase Storage."""
    try: get_supa().storage.from_("pdfs").remove([f"{notice_id}.pdf"])
    except: pass
    try: get_supa().storage.from_("pages").remove([f"{notice_id}/text_map.json"])
    except: pass

def copy_storage(src_id: str, dst_id: str):
    """Copy PDF + text_map from one notice to another."""
    pdf = load_pdf(src_id)
    if pdf:
        store_pdf(dst_id, pdf)
    tm = load_text_map(src_id)
    if tm:
        store_text_map(dst_id, tm)


# ── Authentication ─────────────────────────────────────
async def get_current_user(request: Request):
    """Extract and verify JWT. Returns dict with id, email, org_id, role."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    token = auth[7:]
    try:
        user_resp = get_supa().auth.get_user(token)
        u = user_resp.user
        if not u:
            raise HTTPException(401, "Invalid token")
        # Auto-register in user_roles on first API call (no org yet)
        existing = db_get("user_roles", u.id, id_col="user_id")
        if not existing:
            db_insert("user_roles", {"user_id": u.id, "email": u.email, "role": "uploader"})
            print(f"  [AUTH] New user registered: {u.email} ({u.id})")
            existing = {"user_id": u.id, "email": u.email, "role": "uploader", "org_id": None}
        return {
            "id": u.id,
            "email": u.email,
            "org_id": existing.get("org_id"),
            "role": existing.get("role", "uploader"),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(401, f"Authentication failed: {str(e)[:100]}")

def is_admin(user: dict) -> bool:
    return user.get("role") == "admin"

def require_uploader(user: dict):
    """Raise 403 if viewer."""
    if user.get("role") == "viewer":
        raise HTTPException(403, "열람 전용 계정입니다")

async def require_admin(request: Request):
    """Verify user is admin. Returns user dict."""
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(403, "Admin access required")
    return user

async def check_notice_access(notice_id: str, user: dict):
    """Check if user's org owns the notice (or is admin). Returns notice DB row."""
    r = db_get("notices", notice_id)
    if not r:
        raise HTTPException(404, "Notice not found")
    if is_admin(user):
        return r
    if r.get("org_id") and r["org_id"] != user.get("org_id"):
        raise HTTPException(403, "Access denied")
    return r

def _get_target_org(user: dict, form_org_id: str = None) -> str:
    """Determine which org_id to use for saving data.
    Admin can specify org_id; normal users use their own."""
    if is_admin(user) and form_org_id:
        return form_org_id
    return user.get("org_id") or ""


# ── Duplicate Detection ────────────────────────────────
import re as _re

def _make_fund_id_key(full_name: str) -> str:
    """Generate Fund_ID_Key from full name.
    Keeps critical words (Partners, Fund, Trust, Investment, Capital).
    Removes legal entity suffixes, vehicle/parallel distinguishers.
    Converts Roman numerals to Arabic. Series funds stay separate."""
    if not full_name or full_name == "N/A":
        return ""
    s = full_name
    s = _re.sub(r'\(.*?\)', '', s)           # remove parenthetical
    s = _re.sub(r'"[^"]*"', '', s)           # remove quoted aliases
    s = _re.sub(r'\s*[&]\s*.*', '', s)       # remove & parallel entities
    # Legal suffixes — keep Partners, Fund, Trust, Investment, Capital
    s = _re.sub(r',?\s*\b(L\.?P\.?|LLC|L\.?L\.?C\.?|Ltd\.?|SCSp|SCA|Inc\.?|Corp\.?|Limited)\b', '', s, flags=_re.I)
    # Vehicle/parallel distinguishers — remove so parallel funds merge
    s = _re.sub(r'\b(Parallel|Lux|Main|Feeder|Onshore|Offshore|Co-?Invest|AIV)\b', '', s, flags=_re.I)
    # Roman numerals → Arabic (word-boundary only)
    _roman = [('XVIII','18'),('XVII','17'),('XVI','16'),('XV','15'),('XIV','14'),
              ('XIII','13'),('XII','12'),('XI','11'),('X','10'),('IX','9'),
              ('VIII','8'),('VII','7'),('VI','6'),('V','5'),('IV','4'),
              ('III','3'),('II','2')]
    for roman, arabic in _roman:
        s = _re.sub(r'\b' + roman + r'\b', arabic, s)
    # Remove trailing single-letter vehicle suffix (e.g., "Fund A" → "Fund", "VIII-B" → "8")
    s = _re.sub(r'[,.:;\'"&]', '', s)
    s = _re.sub(r'\s+', ' ', s).strip().lower()
    s = _re.sub(r'[\s-]+[a-c]$', '', s)     # trailing -a, -b, -c or " a", " b", " c"
    s = _re.sub(r'\s+', ' ', s).strip()
    return s


def _make_duplicate_key(header: dict) -> str:
    """Generate a composite key from parsed header for duplicate detection."""
    net = header.get("LP_net_amount")
    try:
        net_rounded = str(round(float(net or 0), 2))
    except (ValueError, TypeError):
        net_rounded = "0"
    fund_key = str(header.get("Fund_ID_Key", "") or header.get("Underlying_Fund_Name_full", "") or "").strip().lower()
    parts = [
        str(header.get("issue_date", "") or "").strip(),
        str(header.get("LP_code", "") or "").strip().lower(),
        fund_key,
        net_rounded,
        str(header.get("notice_type", "") or "").strip().lower(),
    ]
    return hashlib.sha256("|".join(parts).encode()).hexdigest()[:16]


def _find_duplicate(pdf_hash: str, header: dict = None, org_id: str = None):
    """Check for duplicates within the same organization's notices.
    Returns (type, existing_row) or (None, None)."""
    if pdf_hash:
        q = get_supa().table("notices").select("*").eq("pdf_hash", pdf_hash)
        if org_id:
            q = q.eq("org_id", org_id)
        r = q.limit(1).execute()
        if r.data:
            return "exact", r.data[0]
    if header:
        dup_key = _make_duplicate_key(header)
        q = get_supa().table("notices").select("*").eq("duplicate_key", dup_key)
        if org_id:
            q = q.eq("org_id", org_id)
        r = q.limit(1).execute()
        if r.data:
            return "content", r.data[0]
    return None, None


# Temporary store for duplicate notices awaiting user decision
_pending_duplicates: dict = {}  # new_notice_id → full result dict


# ── Smart Page Targeting ───────────────────────────────
_AMOUNT_PATTERN = _re.compile(r'[\d,]+\.\d{2}|\([\d,]+\.\d{2}\)|[\d,]{4,}')
_LI_KEYWORDS = _re.compile(
    r'management\s*fee|capital\s*call|distribution|organizational|'
    r'partnership\s*expense|carried\s*interest|return\s*of\s*capital|'
    r'recallable|withholding|clawback|fund\s*expense|'
    r'partner\s*share|lp\s*amount|investor\s*share',
    _re.I
)

def identify_line_item_pages(text_map: list, page_count: int) -> list:
    """Identify pages most likely to contain the line items table.
    Uses number density, table keywords, and row structure — no AI."""
    if page_count <= 2:
        return list(range(1, page_count + 1))

    scores = {}
    for p in range(1, page_count + 1):
        entries = [t for t in text_map if t["p"] == p]
        if not entries:
            continue
        all_text = " ".join(t["t"] for t in entries)
        score = 0

        # 1. Amount pattern density
        amounts = _AMOUNT_PATTERN.findall(all_text)
        if len(amounts) >= 6:
            score += 3
        elif len(amounts) >= 3:
            score += 1

        # 2. Line item keywords
        kw_hits = len(_LI_KEYWORDS.findall(all_text))
        if kw_hits >= 3:
            score += 3
        elif kw_hits >= 1:
            score += 2

        # 3. Row repetition (table structure)
        y_vals = sorted(set(round(t["y0"]) for t in entries))
        if len(y_vals) >= 8:
            gaps = [y_vals[i + 1] - y_vals[i] for i in range(len(y_vals) - 1)]
            regular_gaps = [g for g in gaps if 8 < g < 25]
            if len(regular_gaps) >= 5:
                score += 2

        if score > 0:
            scores[p] = score

    if not scores:
        return list(range(1, page_count + 1))

    ranked = sorted(scores.items(), key=lambda x: -x[1])
    threshold = max(1, ranked[0][1] * 0.5)
    result = [p for p, s in ranked if s >= threshold][:3]
    print(f"  [PAGE] Line item page scores: {dict(ranked[:5])} → selected: {sorted(result)}")
    return sorted(result) if result else list(range(1, page_count + 1))

# ── PDF Processing (text extraction only — images rendered client-side by PDF.js) ──

# ── Multi-LP (Omnibus) Notice Detection ───────────────
_INVESTOR_ID_PATTERN = _re.compile(r'^\d{3,7}$')

def extract_investor_ids(text_map: list) -> list:
    """Extract Investor ID candidates from text_map.
    Omnibus notices have a schedule table with numeric IDs in the leftmost column."""
    candidates = {}  # id_text → count
    for entry in text_map:
        text = entry.get('t', '').strip()
        if not _INVESTOR_ID_PATTERN.match(text):
            continue
        pw = entry.get('pw', 612)
        x_pct = entry.get('x0', 0) / pw * 100 if pw else 0
        # Investor ID column is typically in the leftmost ~20% of the page
        if x_pct > 22:
            continue
        candidates[text] = candidates.get(text, 0) + 1
    # Sort numerically
    return sorted(candidates.keys(), key=lambda x: int(x))


def _extract_fund_name_preview(text_map: list) -> str:
    """Try to extract fund name from text_map for preview in multi-LP popup.
    Look for common patterns on page 1."""
    page1 = [t for t in text_map if t.get('p') == 1]
    # Concatenate page 1 text by lines
    if not page1:
        return ""
    lines = {}
    for t in page1:
        y_key = round(t['y0'] / 4) * 4
        if y_key not in lines:
            lines[y_key] = []
        lines[y_key].append(t)
    # Look for "KKR ...", "Fund Name ...", etc. in top portion
    fund_kw = _re.compile(r'(fund|partners|investors|infrastructure|capital|trust)', _re.I)
    for y_key in sorted(lines.keys()):
        line_text = " ".join(t['t'] for t in sorted(lines[y_key], key=lambda t: t['x0']))
        if fund_kw.search(line_text) and len(line_text) > 15 and len(line_text) < 120:
            return line_text.strip()
    return ""


def extract_lp_row_data(pdf_bytes: bytes, lp_code: str) -> dict:
    """Pre-extract a specific Investor ID's row data from an omnibus PDF using pdfplumber.
    Used exclusively by parse_multi_lp — never called for single-LP notices.

    Returns dict with:
      found: bool
      page: int (1-based)
      y0: float (pt from top)
      ph: float (page height in pt)
      pw: float (page width in pt)
      row_text_cleaned: str (broken-number-fixed row text)
      numbers: list[str] (extracted numeric values)
      column_headers: str (multi-line header text above the data row)
    """
    import pdfplumber
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for pi, page in enumerate(pdf.pages):
                pw, ph = page.width, page.height

                # 1. Find the word matching lp_code in the left 22% of the page
                words = page.extract_words()
                lp_word = None
                for w in words:
                    if w['text'].strip() == lp_code and w['x0'] / pw * 100 < 22:
                        lp_word = w
                        break
                if not lp_word:
                    continue

                y_top = lp_word['top']
                y_bot = lp_word.get('bottom', y_top + 12)

                # 2. Crop the LP's data row (generous vertical margin)
                row_y0 = max(0, y_top - 3)
                row_y1 = min(ph, y_bot + 5)
                cropped_row = page.crop((0, row_y0, pw, row_y1))
                raw_text = (cropped_row.extract_text() or "").strip()

                # 3. Fix broken numbers — pdfplumber splits digits across narrow gaps
                #    "3 ,929,035" → "3,929,035"   |   "9 6,208" → "96,208"
                #    IMPORTANT: Use negative lookbehind (?<!\d) to only merge when
                #    the left digit is standalone (not part of a multi-digit number).
                #    Without this, "8078 602,404" would wrongly merge to "8078602,404".
                cleaned = raw_text.split("\n")[0]  # Take only the LP's own row (avoid next row)
                cleaned = _re.sub(r'(\d)\s+,', r'\1,', cleaned)            # "3 ,929" → "3,929"
                cleaned = _re.sub(r'\$\s*', '', cleaned)                   # strip $ signs first
                cleaned = _re.sub(r'(?<!\d)(\d)\s+(\d)', r'\1\2', cleaned) # single-digit split only
                cleaned = _re.sub(r'\s+', ' ', cleaned).strip()

                numbers = _re.findall(r'\([\d,]+\.?\d*\)|[\d,]+\.?\d*', cleaned)
                # Remove the LP code itself from the number list
                numbers = [n for n in numbers if n != lp_code]

                # 4. Extract column headers from the area above this data row
                #    Headers are typically 60-100pt above the first data row
                hdr_y0 = max(0, y_top - 110)
                hdr_y1 = max(0, y_top - 8)
                if hdr_y1 > hdr_y0:
                    cropped_hdr = page.crop((0, hdr_y0, pw, hdr_y1))
                    header_text = (cropped_hdr.extract_text() or "").strip()
                    # Also fix broken numbers in headers
                    header_text = _re.sub(r'(\d)\s+,', r'\1,', header_text)
                    header_text = _re.sub(r'(?<!\d)(\d)\s+(\d)', r'\1\2', header_text)
                else:
                    header_text = ""

                print(f"  [LP-ROW] LP {lp_code}: page {pi+1}, y={y_top:.1f}pt, "
                      f"{len(numbers)} values extracted")
                print(f"  [LP-ROW] Cleaned: {cleaned[:120]}...")

                return {
                    "found": True,
                    "page": pi + 1,
                    "y0": y_top,
                    "y1": y_bot,
                    "ph": ph,
                    "pw": pw,
                    "row_text_raw": raw_text,
                    "row_text_cleaned": cleaned,
                    "numbers": numbers,
                    "column_headers": header_text,
                }
    except Exception as e:
        print(f"  [LP-ROW] Error extracting LP {lp_code}: {e}")

    return {"found": False}


def process_pdf_text(pdf_bytes: bytes, notice_id: str) -> dict:
    """Save PDF to Supabase Storage, extract text coordinates with pdfplumber."""
    pdf_hash = hashlib.md5(pdf_bytes).hexdigest()[:12]

    # Save PDF to Supabase Storage
    store_pdf(notice_id, pdf_bytes)

    result = {"pdf_hash": pdf_hash, "page_count": 0, "pages": []}

    # Extract text coordinates (word-level + line-level)
    try:
        import pdfplumber
        text_map = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            result["page_count"] = len(pdf.pages)
            for pi, page in enumerate(pdf.pages):
                pw, ph = round(page.width, 1), round(page.height, 1)
                try:
                    chars = page.chars
                    if not chars:
                        continue
                    chars_sorted = sorted(chars, key=lambda c: (round(c["top"]/2)*2, c["x0"]))
                    lines = []
                    cur_line = [chars_sorted[0]]
                    for c in chars_sorted[1:]:
                        if abs(c["top"] - cur_line[0]["top"]) < 4:
                            cur_line.append(c)
                        else:
                            lines.append(cur_line)
                            cur_line = [c]
                    lines.append(cur_line)
                    for line_chars in lines:
                        line_chars.sort(key=lambda c: c["x0"])
                        words_in_line = []
                        cur_word_chars = [line_chars[0]]
                        for c in line_chars[1:]:
                            gap = c["x0"] - cur_word_chars[-1].get("x1", cur_word_chars[-1]["x0"] + 5)
                            if gap < 3:
                                cur_word_chars.append(c)
                            else:
                                words_in_line.append(cur_word_chars)
                                cur_word_chars = [c]
                        words_in_line.append(cur_word_chars)
                        line_word_entries = []
                        for wchars in words_in_line:
                            word_text = "".join(c.get("text","") for c in wchars).strip()
                            if not word_text:
                                continue
                            entry = {
                                "p": pi + 1, "t": word_text,
                                "x0": round(min(c["x0"] for c in wchars), 1),
                                "y0": round(min(c["top"] for c in wchars), 1),
                                "x1": round(max(c.get("x1", c["x0"]+5) for c in wchars), 1),
                                "y1": round(max(c.get("bottom", c["top"]+10) for c in wchars), 1),
                                "pw": pw, "ph": ph,
                            }
                            if len(word_text) > 1:
                                text_map.append(entry)
                            line_word_entries.append(entry)
                        if line_word_entries:
                            full_line = " ".join(e["t"] for e in line_word_entries)
                            if len(full_line) > 3:
                                text_map.append({
                                    "p": pi + 1, "t": full_line,
                                    "x0": round(min(e["x0"] for e in line_word_entries), 1),
                                    "y0": round(min(e["y0"] for e in line_word_entries), 1),
                                    "x1": round(max(e["x1"] for e in line_word_entries), 1),
                                    "y1": round(max(e["y1"] for e in line_word_entries), 1),
                                    "pw": pw, "ph": ph,
                                })
                    print(f"  [INFO] Page {pi+1}: {len(chars)} chars → {len([t for t in text_map if t['p']==pi+1])} text entries")
                except Exception as page_err:
                    print(f"  [WARN] Page {pi+1} extraction failed: {page_err}")
                    continue
        tm_json = json.dumps(text_map, ensure_ascii=False)
        store_text_map(notice_id, text_map)
        print(f"  [INFO] Text map: {len(text_map)} items for {notice_id}")
    except Exception as e:
        print(f"[WARN] pdfplumber failed: {e}")
        store_text_map(notice_id, [])
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                result["page_count"] = len(pdf.pages)
        except:
            pass

    return result


# ── Gemini API ──────────────────────────────────────────
ANALYSIS_PROMPT = """You are a financial document parser for LP fund notices. Extract ALL data into JSON. Numbers in parentheses like (1,234.56) are NEGATIVE.

Rules:
- If not found use "N/A"
- A dash "-" or em-dash "—" in a number column means zero (0). Do NOT interpret it as a negative sign or missing value.
- Numbers in parentheses like (1,234.56) are NEGATIVE.
- LP_signed_amount: positive=outflow(call), negative=inflow(distribution)
- Commitment_affecting: true/false
- notice_type: net>0→Call, <0→Distribution, 0→Adjustment
- LP_cashDirection: positive→outflow, negative→inflow
- Transaction_type: positive→call, negative→distribution
- LP_signed_amount_source: describe PDF location
- Short names (Underlying_Fund_Name_short): 
  1. Use the GP's commonly known abbreviation if one exists (e.g., "DBP III" for DigitalBridge Partners III, "ACIP" for Asia Climate Infrastructure Partners).
  2. If the PDF itself mentions an abbreviation (e.g., the "Fund", the "Partnership"), use the more specific market-known name instead.
  3. Keep it concise (under 25 chars) but uniquely identifiable.

- LP_Name_short: When the LP name contains a trustee/custodian structure, use ONLY the beneficiary part.
  Pattern: "AAA Bank as Trustee of BBB Trust" → LP_Name_short should be based on "BBB Trust", not "AAA Bank".
  Examples:
    "Standard Chartered Bank as Trustee of PineStreet US Infra Trust No.4" → "PineStreet Trust No.4"
    "Shinhan Bank as Custodian for Korea Pension Fund" → "Korea Pension Fund"
    "AAA f/b/o BBB Capital" → "BBB Capital"
  If no trustee/custodian pattern exists, abbreviate the full LP name normally.

- Fund_ID_Key: A normalized identifier for the fund that must be IDENTICAL across all notices for the same fund. Rules:
  1. Use the FULL fund name (not abbreviation) as the base — if PDF shows both abbreviated and full names, always use the full name.
  2. Remove ALL legal entity suffixes and their preceding punctuation: ", L.P.", " LP", ", LLC", " Ltd.", ", SCSp", ", SCA", " Inc.", " Corp.", " Limited", ", L.L.C."
     Remove the comma/space before the suffix too: "ACIP Parallel Fund A, L.P." → remove ", L.P." → "ACIP Parallel Fund A"
  3. Remove parenthetical clarifications: (the "XX Fund"), ("Parallel Fund"), (collectively, the "Combined Fund")
  4. KEEP critical naming words: Partners, Fund, Trust, Investment, Capital — these distinguish different funds.
  5. Convert Roman numerals to Arabic: I→1, II→2, III→3, IV→4, V→5, VI→6, VII→7, VIII→8, IX→9, X→10, etc.
  6. Lowercase, collapse multiple spaces to single space, trim
  7. The result must be DETERMINISTIC: the same full fund name must ALWAYS produce the same Fund_ID_Key.
  
  CRITICAL — Vehicle / Parallel Fund Grouping:
  8. Different vehicles or parallel structures of the SAME fund must produce the SAME Fund_ID_Key.
     Remove vehicle-distinguishing words: "Parallel", "Lux", "Main", "Feeder", "Onshore", "Offshore", "Co-Invest", "AIV"
     and single-letter vehicle suffixes at the end (A, B, C).
     Example: "ACIP Parallel Fund A, L.P." and "ACIP Fund A, L.P." and "ACIP Lux, SCSp" → ALL become "acip fund"
     Example: "DigitalBridge Partners III Lux, SCSp" and "DigitalBridge Partners III, L.P." → ALL become "digitalbridge partners 3"
  
  CRITICAL — Series Funds are SEPARATE:
  9. Funds with different series numbers (I, II, III, IV, etc.) are DIFFERENT funds and must have DIFFERENT Fund_ID_Keys.
     Example: "KKR Asia Fund III" → "kkr asia fund 3" (different from "kkr asia fund 4")
     Example: "ArcLight Fund VII" vs "ArcLight Fund VIII" → "arclight fund 7" vs "arclight fund 8" (SEPARATE)
  
  Examples:
  "ACIP Parallel Fund A, L.P." → "acip fund"
  "ACIP Fund A, L.P." → "acip fund"
  "DigitalBridge Partners III, L.P." → "digitalbridge partners 3"
  "DigitalBridge Partners III Lux, SCSp" → "digitalbridge partners 3"
  "KKR Asia Pacific Infrastructure Investors SCSp" → "kkr asia pacific infrastructure investors"
  "ArcLight Energy Partners Fund VIII-B, L.P." → "arclight energy partners fund 8"

CRITICAL — is_subtotal classification:
- Set is_subtotal=true for rows that are SUBTOTALS, TOTALS, or NET SUMMARY lines.
- Set is_subtotal=false for actual individual transaction items.

CRITICAL — DO NOT include these as line_items (put them in header fields instead):
- "REMAINING CAPITAL COMMITMENTS" section rows are NOT line items.
- Map to header: Capital Commitment→Commitment_original, Remaining Commitments Prior→Unfunded_prior, Remaining Commitments After→Unfunded_after, Previous Capital Contributions→CumContribPrior.
- Only actual TRANSACTION items belong in line_items.

CRITICAL — Void / Supersede detection:
- If the notice states it voids/supersedes/replaces/amends a prior notice:
  - Set voids_prior_notice to a brief description of what is being voided.
  - Set voids_prior_date to the ISSUE DATE of the prior notice being voided, in "YYYY-MM-DD" format.
    Extract this date from phrases like "previously issued on August 27, 2025" → "2025-08-27",
    "dated March 28, 2025" → "2025-03-28", "notice of March 7, 2025" → "2025-03-07".
    If the prior notice date cannot be determined, set voids_prior_date to null.
- If no prior notice is voided, set both voids_prior_notice and voids_prior_date to null.

CRITICAL — Multi-section table extraction:
- Notice PDFs often contain the same item name (e.g. "Management Fees") in different sections ("Reallocation of Prior Capital Calls", "Current Capital Call", etc.).
- Each occurrence is a SEPARATE line item — extract each one independently.
- Always read the amount from the SAME ROW as the item name. Never substitute a value from a different section for the same item name.
- A dash "-" in the Partner Share / LP Amount column means 0 for that specific row. Do NOT replace it with a value from another section.
- To disambiguate duplicate names, prefix item_name with the section: e.g. "Reallocation - Management Fees" vs "Current Call - Management Fees".
- For LP_signed_amount, always use the column closest to "Partner Share", "LP Amount", or the rightmost amount column.

Return ONLY valid JSON (no markdown/backticks):
{"header":{"notice_number":"","notice_title":"","issue_date":"","due_date":"","LP_Name_full":"","LP_Name_short":"","LP_code":"","Underlying_Fund_Name_full":"","Underlying_Fund_Name_short":"","Fund_ID_Key":"","Underlying_Fund_GP_Name":"","Investment_Class":"LP Interest","Commitment_original":null,"Unfunded_prior":null,"Unfunded_after":null,"Current_Commit_Contribution":null,"Current_Commit_Distribution":null,"CumContribPrior":null,"CumContribAfter":null,"CumDistribPrior":null,"CumDistribAfter":null,"pct_LP_Interest":null,"LP_net_amount":null,"notice_type":"","voids_prior_notice":null,"voids_prior_date":null,"wire_info":[]},"line_items":[{"item_name":"","LP_signed_amount":null,"LP_signed_amount_source":"","LP_absolute_amount":null,"LP_cashDirection":"","Transaction_type":"","Commitment_affecting":false,"is_subtotal":false,"cashOrInkind":"cash","Total_Fund_signed_amount":null,"Total_Fund_absolute_amount":null}]}

wire_info schema — extract ALL wire/payment/banking instructions found in the PDF.
IMPORTANT: Wire instructions have up to 3 distinct entities. Map them carefully:
  - INTERMEDIARY / CORRESPONDENT BANK: The routing bank (often in NY/London). Maps to intermediary_* fields.
  - BENEFICIARY BANK / ACCOUNT WITH INSTITUTION: The recipient's bank. Maps to beneficiary_bank_* fields.
  - BENEFICIARY / FINAL RECIPIENT: The actual account holder. Maps to beneficiary_* fields.
If a PDF only shows 2 levels (e.g., "Bank" + "Account"), put bank info in beneficiary_bank_* and account holder info in beneficiary_*.

CRITICAL — Account Number Isolation Rule:
Each entity's Account Number belongs ONLY to that entity. Do NOT copy or merge account numbers across entities.
  - Account Number under "CORRESPONDENT BANK" heading → intermediary_account_number ONLY
  - Account Number under "BENEFICIARY BANK" heading → beneficiary_bank_account_number ONLY
  - Account Number under "BENEFICIARY" heading → beneficiary_account_number ONLY
  - SWIFT code under "CORRESPONDENT BANK" → intermediary_swift_code ONLY
  - SWIFT code under "BENEFICIARY BANK" → beneficiary_bank_swift_code ONLY
If a field is not present for an entity, leave it as empty string "". Never fill it from another entity's data.

Common heading aliases to watch for:
  - "Correspondent Bank" / "Intermediary Bank" / "Sending Bank" → intermediary_*
  - "Beneficiary Bank" / "Account With Institution" / "Receiving Bank" / "Bank Name" → beneficiary_bank_*
  - "Beneficiary" / "For Further Credit" / "Final Recipient" / "Account Holder" → beneficiary_*

Each element: {"intermediary_bank_name":"","intermediary_bank_address":"","intermediary_swift_code":"","intermediary_account_number":"","beneficiary_bank_name":"","beneficiary_bank_address":"","beneficiary_bank_swift_code":"","beneficiary_bank_aba_routing":"","beneficiary_bank_account_number":"","beneficiary_name":"","beneficiary_account_number":"","reference":"","further_credit":""}
If no wire instructions found, return "wire_info":[] in header. Do NOT infer direction — only extract bank details as written."""


async def call_gemini(pdf_b64: str, model: str = None, temperature: float = None, custom_prompt: str = None, org_id: str = None) -> dict:
    """Call Gemini API with PDF and return parsed result."""
    model = model or _get_setting("gemini_model", GEMINI_MODEL, org_id=org_id)
    if temperature is None:
        temperature = float(_get_setting("gemini_temperature", str(GEMINI_TEMPERATURE), org_id=org_id))
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={GEMINI_KEY}"

    prompt = custom_prompt or ANALYSIS_PROMPT
    payload = {
        "contents": [{"parts": [
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}},
            {"text": prompt}
        ]}],
        "generationConfig": {"temperature": temperature, "maxOutputTokens": 65536}
    }

    async with httpx.AsyncClient(timeout=180.0) as client:
        resp = await client.post(url, json=payload)

    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Gemini API error: {resp.status_code} {resp.text[:300]}")

    data = resp.json()
    usage = data.get("usageMetadata", {})
    candidate = data.get("candidates", [{}])[0]
    raw_text = candidate.get("content", {}).get("parts", [{}])[0].get("text", "")

    if not raw_text.strip():
        raise HTTPException(status_code=502, detail="Gemini returned empty response")

    # Extract JSON
    js = raw_text
    mt_idx = raw_text.find("```")
    if mt_idx >= 0:
        end_idx = raw_text.find("```", mt_idx + 3)
        if end_idx > mt_idx:
            js = raw_text[mt_idx+3:end_idx]
            if js.startswith("json"):
                js = js[4:]
    bs, be = js.find("{"), js.rfind("}")
    if bs == -1 or be == -1 or be <= bs:
        raise HTTPException(status_code=502, detail=f"No JSON in Gemini response: {raw_text[:300]}")
    js = js[bs:be+1]

    try:
        parsed = json.loads(js)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=502, detail=f"JSON parse error: {e}")

    return {
        "parsed": parsed,
        "raw_text": raw_text,
        "tokens": {
            "prompt": usage.get("promptTokenCount", 0),
            "candidates": usage.get("candidatesTokenCount", 0),
            "total": usage.get("totalTokenCount", 0),
        },
        "finish_reason": candidate.get("finishReason", ""),
    }





# ── Post-processing ─────────────────────────────────────
def post_process(parsed: dict) -> tuple:
    """Apply same post-processing as frontend: type normalization, subtotal detection."""
    header = parsed.get("header", {})
    line_items = []
    for it in parsed.get("line_items", []):
        amt = _pn(it.get("LP_signed_amount"))
        tx = "call" if amt and amt > 0 else "distribution" if amt and amt < 0 else (it.get("Transaction_type", "distribution"))
        if tx not in ("call", "distribution"):
            tx = "call" if amt and amt > 0 else "distribution"
        it["Transaction_type"] = tx
        it["LP_cashDirection"] = "outflow" if tx == "call" else "inflow"
        if it.get("LP_absolute_amount") is None and amt is not None:
            it["LP_absolute_amount"] = abs(amt)
        line_items.append(it)

    # Detect subtotals: commitment info rows
    import re
    commit_kw = re.compile(r'^(capital\s*commitment|recallable\s*capital|previous\s*capital\s*contribution|remaining\s*commitment|unfunded\s*commitment|total\s*commitment|current\s*contribution\s*to\s*be|cumulative\s*(contribution|distribution|capital))', re.I)
    for it in line_items:
        name = (it.get("item_name") or "").strip()
        if not it.get("is_subtotal") and commit_kw.match(name):
            it["is_subtotal"] = True

    # notice_type fallback
    if not header.get("notice_type") or header["notice_type"] == "N/A":
        net = _pn(header.get("LP_net_amount"))
        if net is not None:
            header["notice_type"] = "Call" if net > 0 else "Distribution" if net < 0 else "Adjustment"

    # Ensure wire_info exists
    if "wire_info" not in header:
        header["wire_info"] = []

    # Migrate wire_info to v2 (3-entity schema)
    _migrate_header_wire(header)

    # Validate wire entity separation (detect cross-entity contamination)
    _validate_header_wire(header)

    # Normalize: promote lone beneficiary → beneficiary_bank
    _normalize_header_wire_beneficiary(header)

    # Distribution notice: if due_date missing, copy from issue_date
    net = _pn(header.get("LP_net_amount"))
    if net is not None and net < 0:
        due = header.get("due_date")
        issue = header.get("issue_date")
        if (not due or due == "N/A") and issue and issue != "N/A":
            header["due_date"] = issue

    # Fund_ID_Key fallback: if AI didn't provide it, generate from full name
    fid = (header.get("Fund_ID_Key") or "").strip()
    if not fid or fid == "N/A":
        header["Fund_ID_Key"] = _make_fund_id_key(header.get("Underlying_Fund_Name_full", ""))

    # voids_prior_date fallback: extract date from voids_prior_notice text
    vpd = (header.get("voids_prior_date") or "").strip()
    vpn = header.get("voids_prior_notice") or ""
    if vpn and vpn != "N/A" and (not vpd or vpd == "N/A"):
        header["voids_prior_date"] = _extract_date_from_text(str(vpn))

    return header, line_items


_MONTH_MAP = {'january':'01','february':'02','march':'03','april':'04','may':'05','june':'06',
              'july':'07','august':'08','september':'09','october':'10','november':'11','december':'12'}

def _extract_date_from_text(text: str) -> str:
    """Extract the first date from free text. Returns 'YYYY-MM-DD' or ''."""
    import re
    t = text.lower()
    # "Month DD, YYYY" or "Month DD YYYY"
    m = re.search(r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{1,2}),?\s+(\d{4})\b', t)
    if m:
        mm = _MONTH_MAP.get(m.group(1), '')
        if mm:
            return f"{m.group(3)}-{mm}-{m.group(2).zfill(2)}"
    # "DD Month YYYY"
    m = re.search(r'\b(\d{1,2})\s+(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{4})\b', t)
    if m:
        mm = _MONTH_MAP.get(m.group(2), '')
        if mm:
            return f"{m.group(3)}-{mm}-{m.group(1).zfill(2)}"
    # Already ISO "YYYY-MM-DD"
    m = re.search(r'\b(\d{4}-\d{2}-\d{2})\b', t)
    if m:
        return m.group(1)
    return ""


def _pn(v):
    """Parse number, same logic as frontend pN."""
    if v is None or v == "N/A" or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    import re
    c = re.sub(r'[,$\s]', '', str(v))
    c = re.sub(r'\(([^)]+)\)', r'-\1', c)
    try:
        return float(c)
    except:
        return None


# ── PDF Location Mapping ────────────────────────────────
def map_items_to_pdf(line_items: list, notice_id: str, priority_pages: list = None) -> list:
    """
    For each line item, find its location in the PDF by matching
    item name and amount against the text-map. Uses exclusive greedy assignment
    so that each PDF row is claimed by at most one item.
    """
    text_map = load_text_map(notice_id)
    if not text_map:
        print(f"  [WARN] No text_map for {notice_id}, skipping PDF mapping")
        return line_items

    # Section-distinguishing keywords get double weight in name matching
    _SECTION_KW = {'call', 'distribution', 'reallocation', 'current', 'prior',
                   'recallable', 'interest', 'deemed', 'subsequent', 'return'}

    def _collect_candidates(scope_tm, item, name_words):
        """Collect ALL candidate (tm_entry, score) pairs for an item."""
        s_lines = [t for t in scope_tm if len(t.get("t", "")) > 10]
        amt = _pn(item.get("LP_signed_amount"))
        candidates = []  # [(tm_entry, score), ...]
        item_is_negative = (amt is not None and amt < 0)

        # Strategy 1: Match by amount (with sign awareness)
        if amt is not None and amt != 0:
            abs_amt = abs(amt)
            amt_strs = [str(int(abs_amt)), f"{abs_amt:,.0f}", f"{abs_amt:,.2f}",
                        str(abs_amt), str(abs_amt).rstrip('0').rstrip('.')]
            amt_clean = set()
            for a in amt_strs:
                stripped = a.replace(",", "").replace(" ", "")
                amt_clean.add(stripped)
                amt_clean.add(stripped.replace(".", ""))

            amt_matches = []
            for tm in scope_tm:
                raw = tm.get("t", "").strip()
                # Detect sign from original text (parentheses = negative)
                has_parens = ("(" in raw and ")" in raw)
                txt_raw = raw.replace(",", "").replace(" ", "").replace("$", "").replace("(", "").replace(")", "")
                txt_nodot = txt_raw.replace(".", "")
                for ac in amt_clean:
                    if len(ac) > 1 and (txt_raw == ac or txt_nodot == ac):
                        # Fix 1: Sign bonus — matching sign gets +10
                        sign_bonus = 10 if (has_parens == item_is_negative) else 0
                        amt_matches.append((tm, sign_bonus))
                        break

            # Score each candidate with name proximity + section keyword weighting
            for cand, sign_bonus in amt_matches:
                row_texts = [t for t in s_lines if abs(t["y0"] - cand["y0"]) < 5 and t["p"] == cand["p"]]
                row_combined = " ".join(t["t"].lower() for t in row_texts)

                if name_words:
                    # Fix 2: Section keywords get double weight
                    hits = 0
                    for w in name_words:
                        if w in row_combined:
                            hits += 2 if w in _SECTION_KW else 1
                    if hits > 0:
                        score = 100 + sign_bonus + hits  # base 100 + sign + name quality
                        # Extend bbox to cover name text on same row
                        extended = dict(cand)
                        for rt in row_texts:
                            if any(w in rt["t"].lower() for w in name_words):
                                extended["x0"] = min(extended.get("x0", rt["x0"]), rt["x0"])
                                extended["x1"] = max(extended.get("x1", rt["x1"]), rt["x1"])
                        candidates.append((extended, score))
                    else:
                        # Amount matched but no name on row
                        candidates.append((cand, 80 + sign_bonus))
                else:
                    candidates.append((cand, 90 + sign_bonus))

        # Strategy 1b: For amt=0, find rows with dash "-"/"—" near item name keywords
        if amt is not None and amt == 0 and name_words:
            dash_entries = [t for t in scope_tm if t.get("t", "").strip() in ("-", "—", "–", "-")]
            for dash_tm in dash_entries:
                row_texts = [t for t in s_lines if abs(t["y0"] - dash_tm["y0"]) < 5 and t["p"] == dash_tm["p"]]
                row_combined = " ".join(t["t"].lower() for t in row_texts)
                hits = 0
                for w in name_words:
                    if w in row_combined:
                        hits += 2 if w in _SECTION_KW else 1
                if hits > 0:
                    extended = dict(dash_tm)
                    for rt in row_texts:
                        if any(w in rt["t"].lower() for w in name_words):
                            extended["x0"] = min(extended.get("x0", rt["x0"]), rt["x0"])
                            extended["x1"] = max(extended.get("x1", rt["x1"]), rt["x1"])
                    candidates.append((extended, 95 + hits))  # slightly below amount match

        # Strategy 2: Name keywords only (fallback when no amount candidates)
        if not candidates and name_words:
            s_lines = [t for t in scope_tm if len(t.get("t", "")) > 10]
            for tm in s_lines:
                txt_lower = tm.get("t", "").lower()
                hits = 0
                for w in name_words:
                    if w in txt_lower:
                        hits += 2 if w in _SECTION_KW else 1
                if hits > 0:
                    score = min((hits / max(len(name_words), 1)) * 80, 79)
                    candidates.append((tm, score))

        return candidates

    # Build scoped text_maps
    if priority_pages:
        priority_tm = [t for t in text_map if t["p"] in priority_pages]
        fallback_tm = text_map
    else:
        priority_tm = text_map
        fallback_tm = None

    # ── Phase 1: Collect all candidates for every item ──
    all_candidates = []  # [(item_idx, tm_entry, score), ...]
    for i, item in enumerate(line_items):
        name = item.get("item_name", "") or ""
        name_words = [w.lower() for w in name.split() if len(w) > 3]

        cands = _collect_candidates(priority_tm, item, name_words)

        # Fallback to all pages if priority yielded nothing good
        if fallback_tm and fallback_tm is not priority_tm:
            best_priority = max((s for _, s in cands), default=0)
            if best_priority <= 50:
                cands2 = _collect_candidates(fallback_tm, item, name_words)
                cands.extend(cands2)

        for tm_entry, score in cands:
            all_candidates.append((i, tm_entry, score))

    # ── Phase 2: Exclusive greedy assignment ──
    # Sort by score descending — highest confidence matches are assigned first
    all_candidates.sort(key=lambda x: -x[2])

    assigned_items = set()  # item indices already assigned
    assigned_locs = set()   # (page, round(y0)) already claimed

    for item_idx, tm_entry, score in all_candidates:
        if item_idx in assigned_items:
            continue
        if score <= 50:
            continue

        loc_key = (tm_entry["p"], round(tm_entry["y0"]))
        if loc_key in assigned_locs:
            continue  # This row is already claimed by another item

        # Assign
        pw = tm_entry.get("pw", 612)
        ph = tm_entry.get("ph", 792)
        line_items[item_idx]["_pdf_loc"] = {
            "page": tm_entry["p"],
            "x_pct": round(tm_entry["x0"] / pw * 100, 1),
            "y_pct": round(tm_entry["y0"] / ph * 100, 1),
            "w_pct": round((tm_entry["x1"] - tm_entry["x0"]) / pw * 100, 1),
            "h_pct": round((tm_entry["y1"] - tm_entry["y0"]) / ph * 100, 1),
            "match_score": min(round(score), 100),  # cap display at 100
            "matched_text": tm_entry.get("t", "")[:60],
        }
        assigned_items.add(item_idx)
        assigned_locs.add(loc_key)

        name = (line_items[item_idx].get("item_name", "") or "")[:30]
        print(f"    Item '{name}' → page {tm_entry['p']}, y={line_items[item_idx]['_pdf_loc']['y_pct']}% (score:{score:.0f})")

    # Log unmatched items
    for i, item in enumerate(line_items):
        if i not in assigned_items and not item.get("_pdf_loc"):
            item["_pdf_loc"] = None
            name = (item.get("item_name", "") or "")[:30]
            print(f"    Item '{name}' → NOT FOUND in PDF")

    return line_items


# ── Line Item Amount Verification ──────────────────────
def verify_line_item_amounts(line_items: list, text_map: list, priority_pages: list = None) -> list:
    """Cross-check AI-parsed LP_signed_amount against pdfplumber text_map.
    Flags items where the amount doesn't appear on the same row as the item name.
    priority_pages: restrict verification to these pages to avoid false matches."""
    if not line_items or not text_map:
        return line_items

    # If priority pages specified, filter text_map to those pages only
    if priority_pages:
        scoped_tm = [t for t in text_map if t["p"] in priority_pages]
        print(f"  [VERIFY] Scoped to pages {priority_pages}: {len(scoped_tm)}/{len(text_map)} entries")
    else:
        scoped_tm = text_map

    lines = [t for t in scoped_tm if len(t.get("t", "")) > 10]
    words = [t for t in scoped_tm if 1 < len(t.get("t", "")) <= 15]

    def _amt_matches(text, target_abs):
        """Check if a text string represents the target absolute amount."""
        t = text.replace(",", "").replace("$", "").replace("(", "").replace(")", "").replace(" ", "").strip()
        if not t:
            return False
        try:
            val = float(t)
            return abs(val - target_abs) < 0.015
        except:
            pass
        # Also try without decimal
        t_nodot = t.replace(".", "")
        target_strs = {
            str(int(target_abs)),
            f"{target_abs:.0f}",
            f"{target_abs:.2f}".replace(".", ""),
        }
        return t_nodot in target_strs

    for item in line_items:
        if item.get("is_subtotal"):
            continue
        amt = _pn(item.get("LP_signed_amount"))
        name = item.get("item_name", "") or ""
        if amt is None or amt == 0:
            continue  # 0 or N/A — skip verification

        abs_amt = abs(amt)
        name_words = [w.lower() for w in name.split() if len(w) > 3]
        if not name_words:
            continue

        # Find rows in text_map that contain the item name keywords
        name_rows = []  # (page, y0) tuples where name keywords appear
        for tm in lines:
            txt_lower = tm["t"].lower()
            hits = sum(1 for w in name_words if w in txt_lower)
            if hits >= max(1, len(name_words) // 2):
                name_rows.append(tm)

        # Check if the amount appears near any of those name rows
        verified = False
        for nr in name_rows:
            # Get all text entries on the same row (y ± 5pt, same page)
            row_entries = [t for t in scoped_tm
                         if abs(t["y0"] - nr["y0"]) < 5 and t["p"] == nr["p"]]
            for re_entry in row_entries:
                if _amt_matches(re_entry["t"], abs_amt):
                    verified = True
                    break
            if verified:
                break

        if verified:
            item["_amount_verified"] = True
        else:
            item["_amount_verified"] = False
            # Collect candidate amounts from the name rows
            candidates = []
            for nr in name_rows:
                row_entries = [t for t in scoped_tm
                             if abs(t["y0"] - nr["y0"]) < 5 and t["p"] == nr["p"]]
                for re_entry in row_entries:
                    txt = re_entry["t"].replace(",", "").replace("$", "").replace(" ", "").strip()
                    # Check if it looks like a number
                    raw = txt.replace("(", "").replace(")", "")
                    try:
                        val = float(raw)
                        if val != 0 and abs(val - abs_amt) > 0.015:
                            is_neg = "(" in re_entry["t"]
                            candidates.append(round(-val if is_neg else val, 2))
                    except:
                        pass
            # Deduplicate
            if candidates:
                item["_amount_candidates"] = list(set(candidates))[:5]
            print(f"  [AMT WARN] '{name[:40]}': {amt} NOT found near name in text_map. Candidates: {candidates[:3]}")

    return line_items


# ── Wire Info v1→v2 Migration ─────────────────────────
def _migrate_wire_v2(wire: dict) -> dict:
    """Migrate a single wire_info entry from 2-entity (v1) to 3-entity (v2).
    v1 fields: beneficiary_swift_code, beneficiary_aba_routing, beneficiary_account_name
    v2 fields: beneficiary_bank_swift_code, beneficiary_bank_aba_routing, beneficiary_name
    Safe to call on already-migrated data (idempotent)."""
    if not wire or not isinstance(wire, dict):
        return wire
    # Already v2 if beneficiary_bank_swift_code exists with a value
    if wire.get("beneficiary_bank_swift_code"):
        return wire

    new = dict(wire)
    # Move v1 beneficiary_swift_code → beneficiary_bank_swift_code
    if "beneficiary_swift_code" in wire and not new.get("beneficiary_bank_swift_code"):
        new["beneficiary_bank_swift_code"] = wire.get("beneficiary_swift_code", "")
    # Move v1 beneficiary_aba_routing → beneficiary_bank_aba_routing
    if "beneficiary_aba_routing" in wire and not new.get("beneficiary_bank_aba_routing"):
        new["beneficiary_bank_aba_routing"] = wire.get("beneficiary_aba_routing", "")
    # Move v1 beneficiary_account_name → beneficiary_name
    if "beneficiary_account_name" in wire and not new.get("beneficiary_name"):
        new["beneficiary_name"] = wire.get("beneficiary_account_name", "")

    # Ensure new fields exist
    new.setdefault("beneficiary_bank_swift_code", "")
    new.setdefault("beneficiary_bank_aba_routing", "")
    new.setdefault("beneficiary_bank_account_number", "")
    new.setdefault("beneficiary_name", "")
    new.setdefault("intermediary_bank_address", "")

    return new


def _migrate_header_wire(header: dict) -> dict:
    """Apply wire v2 migration to all wire_info entries in a header."""
    wires = header.get("wire_info", [])
    if wires and isinstance(wires, list):
        header["wire_info"] = [_migrate_wire_v2(w) for w in wires]
    return header


def _validate_wire_entities(wire: dict) -> dict:
    """Detect and fix cross-entity account number contamination.
    E.g., if intermediary_account_number == beneficiary_bank_account_number,
    one of them was likely misassigned by AI."""
    if not wire or not isinstance(wire, dict):
        return wire

    # Collect all account numbers with their entity
    acct_fields = [
        ("intermediary_account_number", "intermediary"),
        ("beneficiary_bank_account_number", "beneficiary_bank"),
        ("beneficiary_account_number", "beneficiary"),
    ]
    swift_fields = [
        ("intermediary_swift_code", "intermediary"),
        ("beneficiary_bank_swift_code", "beneficiary_bank"),
    ]

    # Check for duplicate account numbers across entities
    seen_accts = {}
    for field, entity in acct_fields:
        val = (wire.get(field) or "").strip()
        if val and val != "N/A":
            if val in seen_accts:
                prev_entity = seen_accts[val]
                print(f"  [WIRE VALIDATE] Account '{val}' duplicated in "
                      f"{prev_entity} and {entity} — clearing {entity}")
                # Keep the first occurrence, clear the duplicate
                wire[field] = ""
                wire[f"_{field}_cleared_duplicate"] = True
            else:
                seen_accts[val] = entity

    # Check for SWIFT code appearing as account number (common AI mistake)
    all_swifts = set()
    for field, entity in swift_fields:
        val = (wire.get(field) or "").strip()
        if val and val != "N/A":
            all_swifts.add(val.upper())

    for field, entity in acct_fields:
        val = (wire.get(field) or "").strip()
        if val and val.upper() in all_swifts:
            print(f"  [WIRE VALIDATE] '{val}' in {field} looks like a SWIFT code — clearing")
            wire[field] = ""

    return wire


def _validate_header_wire(header: dict) -> dict:
    """Apply wire entity validation to all wire_info entries."""
    wires = header.get("wire_info", [])
    if wires and isinstance(wires, list):
        header["wire_info"] = [_validate_wire_entities(w) for w in wires]
    return header


# ── Wire Beneficiary Normalization ───────────────────────
# GP마다 beneficiary_bank vs beneficiary 필드 사용이 혼재됨.
# beneficiary_bank이 비어있고 beneficiary만 있으면 → beneficiary_bank으로 승격
# 둘 다 있으면 → 그대로 유지 (정상 3-entity 구조)
def _normalize_wire_beneficiary(wire: dict) -> dict:
    """Promote beneficiary fields to beneficiary_bank when bank fields are empty."""
    if not wire or not isinstance(wire, dict):
        return wire

    def _has_val(key):
        v = (wire.get(key) or "").strip()
        return v and v != "N/A"

    # Check if beneficiary_bank has meaningful data
    bank_has_data = any(_has_val(f) for f in [
        "beneficiary_bank_name", "beneficiary_bank_swift_code",
        "beneficiary_bank_aba_routing", "beneficiary_bank_account_number"
    ])
    # Check if beneficiary (non-bank) has meaningful data
    benef_has_data = any(_has_val(f) for f in [
        "beneficiary_name", "beneficiary_account_number"
    ])

    # Only promote if bank is empty and beneficiary has data
    if not bank_has_data and benef_has_data:
        promote_map = {
            "beneficiary_name": "beneficiary_bank_name",
            "beneficiary_account_number": "beneficiary_bank_account_number",
        }
        for src, dst in promote_map.items():
            if _has_val(src) and not _has_val(dst):
                wire[dst] = wire[src]
                wire[src] = ""
                print(f"  [WIRE NORMALIZE] Promoted {src} → {dst}: {wire[dst][:30]}")

        # Also check beneficiary_swift_code (legacy field) → beneficiary_bank_swift_code
        if _has_val("beneficiary_swift_code") and not _has_val("beneficiary_bank_swift_code"):
            wire["beneficiary_bank_swift_code"] = wire["beneficiary_swift_code"]
            wire["beneficiary_swift_code"] = ""
        # Also check beneficiary_aba_routing → beneficiary_bank_aba_routing
        if _has_val("beneficiary_aba_routing") and not _has_val("beneficiary_bank_aba_routing"):
            wire["beneficiary_bank_aba_routing"] = wire["beneficiary_aba_routing"]
            wire["beneficiary_aba_routing"] = ""
        # Promote beneficiary_bank_address if available
        if _has_val("beneficiary_address") and not _has_val("beneficiary_bank_address"):
            wire["beneficiary_bank_address"] = wire.get("beneficiary_address", "")
            wire["beneficiary_address"] = ""

    return wire


def _normalize_header_wire_beneficiary(header: dict) -> dict:
    """Apply beneficiary normalization to all wire_info entries."""
    wires = header.get("wire_info", [])
    if wires and isinstance(wires, list):
        header["wire_info"] = [_normalize_wire_beneficiary(w) for w in wires]
    return header


# ── Wire Info Verification ─────────────────────────────
def verify_wire_info(wire_info: list, text_map: list) -> list:
    """Cross-check AI-parsed wire info against pdfplumber text_map.
    Corrects OCR-like errors (1↔7, 6↔8) in account numbers and codes."""
    if not wire_info or not text_map:
        return wire_info

    all_texts = [t["t"] for t in text_map]
    all_texts_joined = " ".join(all_texts)

    critical_fields = [
        "beneficiary_account_number",
        "beneficiary_bank_account_number",
        "beneficiary_bank_swift_code",
        "beneficiary_bank_aba_routing",
        "intermediary_account_number",
        "intermediary_swift_code",
        # v1 legacy names (in case migration hasn't run yet)
        "beneficiary_swift_code",
        "beneficiary_aba_routing",
    ]

    def _clean(s):
        return s.replace(" ", "").replace("-", "").replace(".", "")

    def _levenshtein(s1, s2):
        if len(s1) < len(s2):
            return _levenshtein(s2, s1)
        if len(s2) == 0:
            return len(s1)
        prev = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            curr = [i + 1]
            for j, c2 in enumerate(s2):
                curr.append(min(curr[j] + 1, prev[j + 1] + 1, prev[j] + (c1 != c2)))
            prev = curr
        return prev[-1]

    for wire in wire_info:
        for field in critical_fields:
            ai_value = (wire.get(field) or "").strip()
            if not ai_value or ai_value == "N/A":
                continue

            ai_clean = _clean(ai_value)

            # Step 1: Exact match in text_map
            if ai_clean in _clean(all_texts_joined):
                wire[f"_{field}_verified"] = True
                continue

            # Also check individual text entries for exact match
            exact_found = False
            for t in all_texts:
                if _clean(t) == ai_clean:
                    exact_found = True
                    break
            if exact_found:
                wire[f"_{field}_verified"] = True
                continue

            # Step 2: Fuzzy match — find similar strings in text_map
            candidates = []
            for t in all_texts:
                t_clean = _clean(t)
                if not t_clean:
                    continue
                # Only compare strings of similar length (±2 chars)
                if abs(len(t_clean) - len(ai_clean)) > 2:
                    continue
                dist = _levenshtein(ai_clean, t_clean)
                if 0 < dist <= 2:
                    candidates.append((t, t_clean, dist))

            if candidates:
                candidates.sort(key=lambda x: x[2])
                tm_value = candidates[0][0]
                dist = candidates[0][2]
                # Keep AI value as primary, flag mismatch for user review
                wire[f"_{field}_mismatch"] = True
                wire[f"_{field}_ai_value"] = ai_value
                wire[f"_{field}_tm_value"] = tm_value
                wire[f"_{field}_distance"] = dist
                # DO NOT overwrite wire[field] — AI value is preserved
                print(f"  [WIRE MISMATCH] {field}: AI='{ai_value}' vs TM='{tm_value}' (dist={dist}) — user review needed")
            else:
                wire[f"_{field}_verified"] = False
                print(f"  [WIRE WARN] {field}: '{ai_value}' not found in text_map — manual check needed")

    return wire_info


# ── API Endpoints ───────────────────────────────────────

class UploadResponse(BaseModel):
    notice_id: str
    header: dict
    line_items: list
    tokens: dict
    page_count: int
    elapsed_ms: int

@app.post("/api/upload")
async def upload_notice(
    request: Request,
    file: UploadFile = File(...),
    model: Optional[str] = Form(None),
    org_id: Optional[str] = Form(None)
):
    """Upload PDF, parse with Gemini, process PDF pages.
    Returns SSE stream with stage events: received → parsing → done/duplicate/error."""
    user = await get_current_user(request)
    require_uploader(user)
    target_org = _get_target_org(user, org_id)
    pdf_bytes = await file.read()
    file_name = file.filename

    async def sse_stream():
        t0 = time.time()
        try:
            # Stage 1: received
            yield f"data: {json.dumps({'stage': 'received', 'fileName': file_name})}\n\n"

            pdf_b64 = base64.b64encode(pdf_bytes).decode()
            pdf_hash = hashlib.md5(pdf_bytes).hexdigest()[:12]

            # Phase 1: exact binary duplicate check (before AI — cost = 0)
            dup_type, dup_row = _find_duplicate(pdf_hash, org_id=target_org)
            if dup_type == "exact":
                existing_header = dup_row["header"] if isinstance(dup_row["header"], dict) else json.loads(dup_row["header"])
                yield f"data: {json.dumps({'stage': 'exact_duplicate', 'existing_id': dup_row['id'], 'existing_file_name': dup_row['file_name'], 'existing_date': existing_header.get('issue_date',''), 'existing_fund': existing_header.get('Underlying_Fund_Name_short',''), 'existing_net': existing_header.get('LP_net_amount')}, ensure_ascii=False)}\n\n"
                return

            # Process PDF text extraction (run in separate thread to unblock event loop)
            notice_id = f"n_{int(time.time()*1000)}_{pdf_hash}"
            pdf_info = await asyncio.to_thread(process_pdf_text, pdf_bytes, notice_id)

            # (pdf_bytes is kept in scope until sse_stream completes — 
            #  it's already saved to disk by process_pdf_text)

            # (PDF.js renders in browser — no server-side image generation needed)

            # Smart Page Targeting: identify line item pages from text_map
            li_pages = None
            tm_data = load_text_map(notice_id)
            if tm_data:
                try:
                    li_pages = identify_line_item_pages(tm_data, pdf_info["page_count"])
                except Exception as e:
                    print(f"  [WARN] Page identification failed: {e}")

            # ── Multi-LP (Omnibus) Notice Detection ──
            # Check BEFORE calling AI to save cost
            investor_ids = extract_investor_ids(tm_data) if tm_data else []
            if len(investor_ids) >= 5:
                fund_preview = _extract_fund_name_preview(tm_data)
                print(f"  [MULTI-LP] Detected {len(investor_ids)} investor IDs: {investor_ids[:5]}...")
                yield f"data: {json.dumps({'stage': 'multi_lp_detect', 'notice_id': notice_id, 'investor_ids': investor_ids, 'fund_preview': fund_preview, 'page_count': pdf_info['page_count']}, ensure_ascii=False)}\n\n"
                return  # Stream ends — user picks LP codes, then calls /api/upload/parse-lp

            # Stage 2: parsing — about to call Gemini AI
            yield f"data: {json.dumps({'stage': 'parsing'})}\n\n"

            # Call Gemini with heartbeat — send periodic SSE events so frontend
            # can distinguish "still waiting for API" from "connection dead"
            gemini_task = asyncio.create_task(call_gemini(pdf_b64, model, org_id=target_org))
            HEARTBEAT_SEC = 10
            hb_count = 0
            while not gemini_task.done():
                try:
                    await asyncio.wait_for(asyncio.shield(gemini_task), timeout=HEARTBEAT_SEC)
                except asyncio.TimeoutError:
                    hb_count += 1
                    yield f"data: {json.dumps({'stage': 'heartbeat', 'elapsed': hb_count * HEARTBEAT_SEC})}\n\n"
                except Exception:
                    break  # actual error — will be raised below
            gemini_result = gemini_task.result()  # raises if task failed

            # Free base64 string (~33% larger than PDF) to reduce memory pressure
            del pdf_b64

            # Post-process
            header, line_items = post_process(gemini_result["parsed"])

            # Phase 3: content-based duplicate check (after AI parsing)
            dup_key = _make_duplicate_key(header)
            dup_type2, dup_row2 = _find_duplicate(None, header, org_id=target_org)
            if dup_type2 == "content":
                # AI already spent — save result to pending, let user decide
                elapsed = int((time.time() - t0) * 1000)

                # Still do verification so the pending result is complete
                print(f"  [INFO] Mapping {len(line_items)} items to PDF (pages: {li_pages})...")
                line_items = map_items_to_pdf(line_items, notice_id, priority_pages=li_pages)
                if tm_data:
                    line_items = verify_line_item_amounts(line_items, tm_data, priority_pages=li_pages)
                if header.get("wire_info") and tm_data:
                    header["wire_info"] = verify_wire_info(header["wire_info"], tm_data)

                _pending_duplicates[notice_id] = {
                    "notice_id": notice_id, "file_name": file_name,
                    "header": header, "line_items": line_items,
                    "raw_text": gemini_result["raw_text"],
                    "tokens": gemini_result["tokens"],
                    "pdf_hash": pdf_hash, "page_count": pdf_info["page_count"],
                    "duplicate_key": dup_key, "elapsed_ms": elapsed,
                }
                existing_header = dup_row2["header"] if isinstance(dup_row2["header"], dict) else json.loads(dup_row2["header"])
                yield f"data: {json.dumps({'stage': 'duplicate', 'new_notice_id': notice_id, 'existing_id': dup_row2['id'], 'existing_file_name': dup_row2['file_name'], 'existing_date': existing_header.get('issue_date',''), 'existing_fund': existing_header.get('Underlying_Fund_Name_short',''), 'existing_net': existing_header.get('LP_net_amount'), 'header': header, 'line_items': line_items, 'tokens': gemini_result['tokens'], 'page_count': pdf_info['page_count'], 'elapsed_ms': elapsed}, ensure_ascii=False)}\n\n"
                return

            # No duplicate — normal flow
            print(f"  [INFO] Mapping {len(line_items)} items to PDF (pages: {li_pages})...")
            line_items = map_items_to_pdf(line_items, notice_id, priority_pages=li_pages)

            # Verify wire info
            if header.get("wire_info") and tm_data:
                try:
                    header["wire_info"] = verify_wire_info(header["wire_info"], tm_data)
                except Exception as e:
                    print(f"  [WARN] Wire verification failed: {e}")

            # Verify line item amounts
            if tm_data:
                try:
                    line_items = verify_line_item_amounts(line_items, tm_data, priority_pages=li_pages)
                except Exception as e:
                    print(f"  [WARN] Amount verification failed: {e}")

            # Save to DB
            db_insert("notices", {
                    "id": notice_id, "file_name": file_name, "user_id": user["id"], "org_id": target_org,
                    "header": header, "line_items": line_items,
                    "raw_ai_response": gemini_result["raw_text"],
                    "pdf_hash": pdf_hash, "page_count": pdf_info["page_count"], "duplicate_key": dup_key
                })

            elapsed = int((time.time() - t0) * 1000)

            result = {
                "stage": "done",
                "notice_id": notice_id,
                "header": header,
                "line_items": line_items,
                "tokens": gemini_result["tokens"],
                "page_count": pdf_info["page_count"],
                "elapsed_ms": elapsed,
            }
            yield f"data: {json.dumps(result, ensure_ascii=False)}\n\n"

        except Exception as e:
            import traceback; traceback.print_exc()
            yield f"data: {json.dumps({'stage': 'error', 'detail': str(e)[:500]})}\n\n"

    return StreamingResponse(sse_stream(), media_type="text/event-stream")


# ── Duplicate Resolution ───────────────────────────────

@app.put("/api/notices/{new_id}/resolve-duplicate")
async def resolve_duplicate(new_id: str, body: dict, request: Request):
    user = await get_current_user(request)
    require_uploader(user)
    """Resolve a duplicate notice. action: 'keep_existing' or 'replace_with_new'."""
    action = body.get("action")
    existing_id = body.get("existing_id")
    pending = _pending_duplicates.pop(new_id, None)

    if action == "keep_existing":
        # Discard the new upload — clean up files
        delete_storage(new_id)
        return {"ok": True, "action": "kept_existing", "existing_id": existing_id}

    elif action == "replace_with_new" and pending:
        # Delete existing, save new
        if existing_id:
            db_delete("notices", existing_id)
            delete_storage(existing_id)

        # Save pending as new notice
        db_insert("notices", {
                "id": pending["notice_id"], "file_name": pending["file_name"],
                "user_id": user["id"], "org_id": user.get("org_id", ""),
                "header": pending["header"], "line_items": pending["line_items"],
                "raw_ai_response": pending["raw_text"],
                "pdf_hash": pending["pdf_hash"], "page_count": pending["page_count"],
                "duplicate_key": pending["duplicate_key"]
            })

        return {
            "ok": True, "action": "replaced",
            "deleted_id": existing_id, "new_id": pending["notice_id"],
            "header": pending["header"], "line_items": pending["line_items"],
            "tokens": pending["tokens"], "page_count": pending["page_count"],
        }

    raise HTTPException(400, "Invalid action or missing pending data")


@app.post("/api/notices/{notice_id}/reparse")
async def reparse_notice(notice_id: str, request: Request, model: Optional[str] = Query(None)):
    user = await get_current_user(request)
    require_uploader(user)
    await check_notice_access(notice_id, user)
    # Find existing notice and its PDF
    r = db_get("notices", notice_id)
    if not r:
        raise HTTPException(404, "Notice not found in database")
    target_org = r.get("org_id") or user.get("org_id", "")
    file_name = r["file_name"]
    pdf_bytes = load_pdf(notice_id)
    if not pdf_bytes:
        raise HTTPException(404, f"PDF file not found in storage: {notice_id}")

    async def sse_stream():
        t0 = time.time()
        try:
            yield f"data: {json.dumps({'stage': 'received', 'fileName': file_name})}\n\n"

            pdf_b64 = base64.b64encode(pdf_bytes).decode()
            pdf_hash = hashlib.md5(pdf_bytes).hexdigest()[:12]

            # Re-process text extraction (overwrites existing text_map — same notice_id)
            pdf_info = await asyncio.to_thread(process_pdf_text, pdf_bytes, notice_id)

            # Load text_map
            tm_data = load_text_map(notice_id)
            li_pages = None
            if tm_data:
                try:
                    pass  # tm_data already loaded
                    li_pages = identify_line_item_pages(tm_data, pdf_info["page_count"])
                except Exception as e:
                    print(f"  [WARN] Page identification failed: {e}")

            # Multi-LP check
            investor_ids = extract_investor_ids(tm_data) if tm_data else []
            if len(investor_ids) >= 5:
                fund_preview = _extract_fund_name_preview(tm_data)
                yield f"data: {json.dumps({'stage': 'multi_lp_detect', 'notice_id': notice_id, 'investor_ids': investor_ids, 'fund_preview': fund_preview, 'page_count': pdf_info['page_count']}, ensure_ascii=False)}\n\n"
                return

            yield f"data: {json.dumps({'stage': 'parsing'})}\n\n"

            gemini_task = asyncio.create_task(call_gemini(pdf_b64, model, org_id=target_org))
            hb_count = 0
            while not gemini_task.done():
                try:
                    await asyncio.wait_for(asyncio.shield(gemini_task), timeout=10)
                except asyncio.TimeoutError:
                    hb_count += 1
                    yield f"data: {json.dumps({'stage': 'heartbeat', 'elapsed': hb_count * 10})}\n\n"
                except Exception:
                    break
            gemini_result = gemini_task.result()
            del pdf_b64
            header, line_items = post_process(gemini_result["parsed"])

            # Verification
            line_items = map_items_to_pdf(line_items, notice_id, priority_pages=li_pages)
            if header.get("wire_info") and tm_data:
                try: header["wire_info"] = verify_wire_info(header["wire_info"], tm_data)
                except Exception as e: print(f"  [WARN] Wire verification failed: {e}")
            if tm_data:
                try: line_items = verify_line_item_amounts(line_items, tm_data, priority_pages=li_pages)
                except Exception as e: print(f"  [WARN] Amount verification failed: {e}")

            # Update existing notice in-place (same ID — no orphaned files)
            dup_key = _make_duplicate_key(header)
            db_update("notices", {
                    "header": header, "line_items": line_items,
                    "raw_ai_response": gemini_result["raw_text"],
                    "pdf_hash": pdf_hash, "page_count": pdf_info["page_count"],
                    "duplicate_key": dup_key
                }, notice_id)

            elapsed = int((time.time() - t0) * 1000)
            yield f"data: {json.dumps({'stage': 'done', 'notice_id': notice_id, 'header': header, 'line_items': line_items, 'tokens': gemini_result['tokens'], 'page_count': pdf_info['page_count'], 'elapsed_ms': elapsed}, ensure_ascii=False)}\n\n"

        except Exception as e:
            import traceback; traceback.print_exc()
            yield f"data: {json.dumps({'stage': 'error', 'detail': str(e)[:500]})}\n\n"

    return StreamingResponse(sse_stream(), media_type="text/event-stream")


# ── Multi-LP Targeted Parsing ────────────────────────────

_TARGETED_LP_SUFFIX = """

CRITICAL — TARGETED LP EXTRACTION:
This is a multi-LP omnibus notice containing a schedule with data for many Investor IDs.
Extract data ONLY for Investor ID: {lp_code}

Rules for targeted extraction:
1. LP_code MUST be "{lp_code}".
2. LP_net_amount: Use the "Total Contributions" or "Total Distribution" column value 
   from the row matching Investor ID {lp_code}. This is the ONLY amount that matters.
3. line_items: Extract each column amount from the Investor ID {lp_code} row as separate 
   line items (e.g. "Investment Amount", "Interest Expense", "Management Fee", "Fund Expenses").
   Use column headers as item_name. Use the Investor ID {lp_code} row values ONLY.
4. Commitment fields: Use "Beginning Unused Commitment" and "Ending Unused Commitment" 
   from the Investor ID {lp_code} row for Unfunded_prior and Unfunded_after.
5. Do NOT use "TOTAL LP", "Total", or any summary/subtotal row values.
6. If Investor ID {lp_code} shows a dash "-" or blank for an amount, report 0.
7. LP_Name_full / LP_Name_short: Set to "N/A" (omnibus notices don't name individual LPs).
8. Fund-level info (fund name, dates, wire instructions): extract from the cover letter pages as usual.
"""

_TARGETED_LP_SUFFIX_V2 = """

CRITICAL — TARGETED LP EXTRACTION (PRE-EXTRACTED DATA):
This is a multi-LP omnibus notice. The data for Investor ID {lp_code} has been 
pre-extracted from the PDF schedule table using coordinate-based text extraction.
Do NOT re-read the schedule table from the PDF pages — use ONLY the pre-extracted data below.

=== COLUMN HEADERS (from the table header area above this row) ===
{column_headers}

=== DATA ROW for Investor ID {lp_code} ===
{row_text}

=== NUMERIC VALUES in left-to-right order (LP code excluded) ===
{numbers}

Instructions:
1. LP_code MUST be "{lp_code}".
2. Match each numeric value above to the corresponding column header, left to right.
   - The first large number is typically "Beginning Unused Commitment" → Unfunded_prior.
   - Intermediate columns are investment amounts, fees, expenses → line_items.
   - "Total Contributions" or "Total Distribution" → LP_net_amount.
   - The last commitment number is "Ending Unused Commitment" → Unfunded_after.
3. Each intermediate column = one line_item. Use column headers as item_name.
   Prefix with the section name if visible in headers (e.g. "India Grid Trust - Investment Amount").
4. Numbers in parentheses like (1,234) are NEGATIVE.
5. A dash "-" or blank = 0.
6. LP_Name_full / LP_Name_short: Set to "N/A" (omnibus notices don't name individual LPs).
7. Fund name, dates, wire instructions: extract from the cover letter (page 1) of the PDF as usual.
8. TRUST THE PRE-EXTRACTED NUMBERS. They are taken directly from the PDF text layer.
"""

@app.post("/api/upload/parse-lp")
async def parse_multi_lp(
    request: Request,
    notice_id: str = Form(...),
    lp_codes: str = Form(...),
    model: Optional[str] = Form(None)
):
    """Parse a multi-LP omnibus notice for specific LP codes.
    Returns SSE stream: lp_parsing → lp_done/lp_error per LP → all_done."""
    user = await get_current_user(request)
    require_uploader(user)
    target_org = user.get("org_id", "")
    lp_list = json.loads(lp_codes)
    if not lp_list:
        raise HTTPException(400, "No LP codes provided")

    pdf_bytes = load_pdf(notice_id)
    if not pdf_bytes:
        raise HTTPException(404, f"PDF not found for {notice_id}")

    pdf_b64 = base64.b64encode(pdf_bytes).decode()
    pdf_hash = hashlib.md5(pdf_bytes).hexdigest()[:12]

    # Load text_map (already generated during upload)
    tm_data = load_text_map(notice_id)
    # Get page count
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page_count = len(pdf.pages)
    except:
        page_count = 1

    li_pages = identify_line_item_pages(tm_data, page_count) if tm_data else None

    async def sse_stream():
        t0 = time.time()
        completed = []
        for i, lp_code in enumerate(lp_list):
            lp_code = str(lp_code).strip()
            sub_id = f"{notice_id}_lp{lp_code}"
            try:
                yield f"data: {json.dumps({'stage': 'lp_parsing', 'lp_code': lp_code, 'progress': f'{i+1}/{len(lp_list)}'})}\n\n"

                # ── Pre-extract LP row data from PDF using pdfplumber ──
                lp_row = await asyncio.to_thread(extract_lp_row_data, pdf_bytes, lp_code)

                # ── Build prompt: enhanced (pre-extracted) or fallback (visual) ──
                if lp_row["found"]:
                    numbers_str = ", ".join(lp_row["numbers"])
                    suffix = _TARGETED_LP_SUFFIX_V2 \
                        .replace("{lp_code}", lp_code) \
                        .replace("{column_headers}", lp_row["column_headers"]) \
                        .replace("{row_text}", lp_row["row_text_cleaned"]) \
                        .replace("{numbers}", numbers_str)
                    targeted_prompt = ANALYSIS_PROMPT + suffix
                    print(f"  [MULTI-LP] LP {lp_code}: using pre-extracted data "
                          f"(page {lp_row['page']}, {len(lp_row['numbers'])} values)")
                else:
                    # Fallback: let AI read the table visually (old behavior)
                    targeted_prompt = ANALYSIS_PROMPT + _TARGETED_LP_SUFFIX.replace("{lp_code}", lp_code)
                    print(f"  [MULTI-LP] LP {lp_code}: row not found in text_map, "
                          f"using visual fallback prompt")

                # ── Call Gemini with heartbeat ──
                gemini_task = asyncio.create_task(call_gemini(pdf_b64, model, custom_prompt=targeted_prompt, org_id=target_org))
                hb_count = 0
                while not gemini_task.done():
                    try:
                        await asyncio.wait_for(asyncio.shield(gemini_task), timeout=10)
                    except asyncio.TimeoutError:
                        hb_count += 1
                        yield f"data: {json.dumps({'stage': 'heartbeat', 'lp_code': lp_code, 'elapsed': hb_count * 10})}\n\n"
                    except Exception:
                        break
                gemini_result = gemini_task.result()
                header, line_items = post_process(gemini_result["parsed"])

                # Force LP_code
                header["LP_code"] = lp_code

                # Copy PDF & text_map for sub-notice
                copy_storage(notice_id, sub_id)

                # ── Verification & PDF location mapping — scoped to LP row ──
                if lp_row["found"] and tm_data:
                    lp_page = lp_row["page"]
                    lp_y0 = lp_row["y0"]
                    lp_y1 = lp_row.get("y1", lp_y0 + 12)
                    lp_ph = lp_row.get("ph", 792)
                    lp_pw = lp_row.get("pw", 612)

                    # Run map_items_to_pdf with full text_map but restricted to LP page
                    try:
                        line_items = map_items_to_pdf(line_items, sub_id, priority_pages=[lp_page])
                    except Exception as e:
                        print(f"  [WARN] map_items_to_pdf failed for LP {lp_code}: {e}")

                    # Override _pdf_loc for all non-subtotal items to point to the LP row
                    # This ensures highlights land on the correct row, not a random
                    # row that happens to share the same amount value
                    y_pct = round(lp_y0 / lp_ph * 100, 1)
                    h_pct = round((lp_y1 - lp_y0) / lp_ph * 100, 1)
                    for item in line_items:
                        if item.get("is_subtotal"):
                            continue
                        item["_pdf_loc"] = {
                            "page": lp_page,
                            "x_pct": 1.0,
                            "y_pct": y_pct,
                            "w_pct": 95.0,
                            "h_pct": max(1.2, h_pct),
                            "match_score": 95,
                            "matched_text": f"LP {lp_code} row (pre-extracted)",
                        }

                    # ── Direct amount verification against pre-extracted numbers ──
                    # Standard verify_line_item_amounts fails for omnibus because it
                    # requires item_name keywords and amount on the SAME text row.
                    # In omnibus tables, item_name = column header (y~150) while
                    # amounts are in the data row (y~460) — they're never co-located.
                    # Instead, directly check if each AI-parsed amount exists in the
                    # pre-extracted number list from pdfplumber.
                    pre_amounts = set()
                    for n_str in lp_row["numbers"]:
                        raw = n_str.replace(",", "")
                        # Handle parenthesized negatives: "(1234)" → 1234
                        if raw.startswith("(") and raw.endswith(")"):
                            raw = raw[1:-1]
                        try:
                            pre_amounts.add(round(abs(float(raw)), 2))
                        except (ValueError, TypeError):
                            pass

                    verified_count = 0
                    for item in line_items:
                        if item.get("is_subtotal"):
                            continue
                        amt = _pn(item.get("LP_signed_amount"))
                        if amt is None:
                            continue
                        abs_amt = round(abs(amt), 2)
                        if abs_amt == 0 or abs_amt in pre_amounts:
                            item["_amount_verified"] = True
                            verified_count += 1
                        else:
                            item["_amount_verified"] = False
                            # Check if close match exists (rounding tolerance)
                            close = [a for a in pre_amounts if abs(a - abs_amt) < 1.0]
                            if close:
                                item["_amount_verified"] = True
                                verified_count += 1
                                print(f"  [VERIFY] '{item.get('item_name','')[:30]}': "
                                      f"{abs_amt} ≈ {close[0]} (close match)")
                            else:
                                item["_amount_candidates"] = sorted(pre_amounts)[:5]
                                print(f"  [AMT WARN] '{item.get('item_name','')[:30]}': "
                                      f"{amt} not in pre-extracted numbers")
                    non_sub = [it for it in line_items if not it.get("is_subtotal")]
                    print(f"  [VERIFY] LP {lp_code}: {verified_count}/{len(non_sub)} "
                          f"items verified against pre-extracted numbers")
                else:
                    # Fallback: use original unscoped verification
                    try:
                        line_items = map_items_to_pdf(line_items, sub_id, priority_pages=li_pages)
                    except Exception as e:
                        print(f"  [WARN] map_items_to_pdf failed for LP {lp_code}: {e}")
                    if tm_data:
                        try:
                            line_items = verify_line_item_amounts(line_items, tm_data, priority_pages=li_pages)
                        except Exception as e:
                            print(f"  [WARN] verify_amounts failed for LP {lp_code}: {e}")

                # Wire info verification (always uses full text_map — wire info is on cover pages)
                if header.get("wire_info") and tm_data:
                    try:
                        header["wire_info"] = verify_wire_info(header["wire_info"], tm_data)
                    except Exception as e:
                        print(f"  [WARN] verify_wire failed for LP {lp_code}: {e}")

                # ── Content-based duplicate check before saving ──
                # Without this, re-uploading the same omnibus PDF creates duplicate
                # sub-notices because sub_id contains a timestamp-based notice_id.
                dup_key = _make_duplicate_key(header)
                is_dup = False
                _dup_q = get_supa().table("notices").select("*").eq("duplicate_key", dup_key).eq("org_id", target_org).limit(1).execute()
                existing = _dup_q.data[0] if _dup_q.data else None
                if existing and existing["id"] != sub_id:
                    is_dup = True
                    print(f"  [DUP] LP {lp_code}: content duplicate of "
                          f"'{existing['file_name']}' ({existing['id']}), skipping")
                if is_dup:
                    yield f"data: {json.dumps({'stage': 'lp_done', 'lp_code': lp_code, 'notice_id': existing['id'], 'header': header, 'line_items': line_items, 'tokens': gemini_result['tokens'], 'page_count': page_count, 'progress': f'{i+1}/{len(lp_list)}', 'elapsed_ms': int((time.time()-t0)*1000), 'skipped_duplicate': True}, ensure_ascii=False)}\n\n"
                    completed.append(existing["id"])
                    # Clean up copied files for this sub_id since we're not saving
                    delete_storage(sub_id)
                    continue  # Skip to next LP code

                # Save to DB
                file_label = f"{notice_id}_LP{lp_code}.pdf"
                try: db_delete("notices", sub_id)
                except: pass
                db_insert("notices", {
                    "id": sub_id, "file_name": file_label, "user_id": user["id"], "org_id": target_org,
                    "header": header, "line_items": line_items,
                    "raw_ai_response": gemini_result["raw_text"],
                    "pdf_hash": pdf_hash, "page_count": page_count, "duplicate_key": dup_key
                })

                elapsed = int((time.time() - t0) * 1000)
                yield f"data: {json.dumps({'stage': 'lp_done', 'lp_code': lp_code, 'notice_id': sub_id, 'header': header, 'line_items': line_items, 'tokens': gemini_result['tokens'], 'page_count': page_count, 'progress': f'{i+1}/{len(lp_list)}', 'elapsed_ms': elapsed}, ensure_ascii=False)}\n\n"
                completed.append(sub_id)

            except Exception as e:
                import traceback; traceback.print_exc()
                yield f"data: {json.dumps({'stage': 'lp_error', 'lp_code': lp_code, 'detail': str(e)[:500]})}\n\n"

            # Small delay between LP calls to avoid rate limiting
            if i < len(lp_list) - 1:
                await asyncio.sleep(1)

        yield f"data: {json.dumps({'stage': 'all_done', 'notice_ids': completed})}\n\n"

    return StreamingResponse(sse_stream(), media_type="text/event-stream")


# ── Deferred Multi-LP Notice (LP code unknown) ────────

@app.post("/api/notices/defer-lp")
async def defer_multi_lp(body: dict, request: Request):
    user = await get_current_user(request)
    require_uploader(user)
    """Save a placeholder notice for a multi-LP PDF when the user doesn't know
    the LP code yet. The PDF is already on disk from the upload stage.
    The user can later click this notice and enter their LP code to trigger parsing."""
    _target_org = user.get("org_id", "")
    notice_id = body.get("notice_id", "")
    file_name = body.get("file_name", "")
    fund_preview = body.get("fund_preview", "")
    investor_ids = body.get("investor_ids", [])
    page_count = body.get("page_count", 0)

    # Verify PDF exists in storage
    if not load_pdf(notice_id):
        raise HTTPException(404, f"PDF not found for {notice_id}")

    # Build placeholder header
    header = {
        "notice_type": "Pending",
        "issue_date": "N/A",
        "due_date": "N/A",
        "LP_Name_full": "N/A",
        "LP_Name_short": "N/A",
        "LP_code": "N/A",
        "Underlying_Fund_Name_full": fund_preview or "N/A",
        "Underlying_Fund_Name_short": fund_preview[:25] if fund_preview else "N/A",
        "Fund_ID_Key": _make_fund_id_key(fund_preview) if fund_preview else "",
        "LP_net_amount": None,
        "wire_info": [],
        "_pending_lp": True,
        "_investor_ids": investor_ids,
        "_fund_preview": fund_preview,
    }

    # Save to DB
    _pdf_for_hash = load_pdf(notice_id)
    pdf_hash = hashlib.md5(_pdf_for_hash[:4096]).hexdigest()[:12] if _pdf_for_hash else ""
    db_upsert("notices", {
        "id": notice_id, "file_name": file_name, "user_id": user["id"], "org_id": _target_org,
        "header": header, "line_items": [],
        "raw_ai_response": "", "pdf_hash": pdf_hash,
        "page_count": page_count, "duplicate_key": ""
    })

    print(f"  [DEFER] Saved placeholder notice {notice_id}: {file_name} "
          f"({len(investor_ids)} investor IDs)")

    return {
        "ok": True,
        "notice_id": notice_id,
        "header": header,
        "page_count": page_count,
    }


@app.get("/api/notices")
async def list_notices(request: Request, org_id: Optional[str] = Query(None)):
    user = await get_current_user(request)
    # Admin can filter by org_id
    if is_admin(user):
        if org_id and org_id != "all":
            rows = db_list("notices", order_col="created_at", order_desc=True, org_id=org_id)
        else:
            rows = db_list("notices", order_col="created_at", order_desc=True)
    else:
        rows = db_list("notices", order_col="created_at", order_desc=True, org_id=user["org_id"])
    result = []
    for r in rows:
        header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
        _migrate_header_wire(header)
        _normalize_header_wire_beneficiary(header)
        line_items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")
        result.append({
            "id": r["id"], "fileName": r.get("file_name",""), "analyzedAt": r.get("analyzed_at"),
            "header": header, "lineItems": line_items,
            "is_voided": bool(r.get("is_voided")), "voided_by": r.get("voided_by"),
            "page_count": r.get("page_count") or 0,
            "org_id": r.get("org_id", ""),
        })
    return result


@app.get("/api/notices/{notice_id}")
async def get_notice(notice_id: str, request: Request):
    user = await get_current_user(request)
    r = await check_notice_access(notice_id, user)
    if not r:
        raise HTTPException(404, "Notice not found")
    header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
    _migrate_header_wire(header)
    _normalize_header_wire_beneficiary(header)
    line_items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")
    return {
        "id": r["id"], "fileName": r.get("file_name",""), "analyzedAt": r.get("analyzed_at"),
        "header": header, "lineItems": line_items,
        "rawAiResponse": r.get("raw_ai_response",""),
        "is_voided": bool(r.get("is_voided")), "voided_by": r.get("voided_by"),
        "page_count": r.get("page_count") or 0,
    }


@app.delete("/api/notices/{notice_id}")
async def delete_notice(notice_id: str, request: Request):
    user = await get_current_user(request)
    require_uploader(user)
    await check_notice_access(notice_id, user)
    db_delete("notices", notice_id)
    delete_storage(notice_id)
    return {"ok": True}


@app.put("/api/notices/{notice_id}/items/{item_idx}")
async def update_item(notice_id: str, item_idx: int, updates: dict, request: Request):
    """Update a single line item (toggle type, commit, etc.)."""
    user = await get_current_user(request)
    require_uploader(user)
    r = await check_notice_access(notice_id, user)
    if not r: raise HTTPException(404)
    items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")
    if item_idx < 0 or item_idx >= len(items): raise HTTPException(400, "Invalid index")
    items[item_idx].update(updates)
    db_update("notices", {"line_items": items}, notice_id)
    return {"ok": True, "item": items[item_idx]}


@app.put("/api/notices/{notice_id}/header")
async def update_header(notice_id: str, updates: dict, request: Request):
    """Update header fields. Also handles is_voided/voided_by (DB columns)."""
    user = await get_current_user(request)
    require_uploader(user)
    r = await check_notice_access(notice_id, user)
    if not r: raise HTTPException(404)
    header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
    set_voided = updates.pop("is_voided", None)
    set_voided_by = updates.pop("voided_by", None)
    header.update(updates)
    update_data = {"header": header}
    if set_voided is not None:
        update_data["is_voided"] = bool(set_voided)
        update_data["voided_by"] = set_voided_by
    db_update("notices", update_data, notice_id)
    return {"ok": True}


@app.put("/api/notices/{notice_id}/items")
async def bulk_update_items(notice_id: str, request: Request, items: list = Body(...)):
    user = await get_current_user(request)
    require_uploader(user)
    await check_notice_access(notice_id, user)
    """Replace all line items (for correction apply)."""
    db_update("notices", {"line_items": items}, notice_id)
    return {"ok": True}


# ── PDF Page Images — removed (PDF.js renders client-side) ──
# GET /api/notices/{id}/pdf is used instead


@app.get("/api/notices/{notice_id}/text-map")
async def get_text_map(notice_id: str, request: Request, page: Optional[int] = None):
    user = await get_current_user(request)
    await check_notice_access(notice_id, user)
    text_map = load_text_map(notice_id)
    if page:
        text_map = [t for t in text_map if t["p"] == page]
    return text_map


@app.get("/api/notices/{notice_id}/pdf")
async def get_pdf(notice_id: str, request: Request):
    user = await get_current_user(request)
    await check_notice_access(notice_id, user)
    pdf_bytes = load_pdf(notice_id)
    if not pdf_bytes:
        raise HTTPException(404, "PDF not found")
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={notice_id}.pdf"})


@app.get("/api/notices/{notice_id}/pdf-url")
async def get_pdf_url(notice_id: str, request: Request):
    """Return a Supabase Storage signed URL for direct PDF download.
    Bypasses server as PDF relay — browser downloads directly from Supabase CDN."""
    user = await get_current_user(request)
    await check_notice_access(notice_id, user)
    try:
        result = get_supa().storage.from_("pdfs").create_signed_url(
            f"{notice_id}.pdf", 300)  # 5 minutes
        url = result.get("signedURL") or result.get("signedUrl") or ""
        if not url:
            raise HTTPException(500, "Failed to create signed URL")
        return {"url": url}
    except HTTPException:
        raise
    except Exception as e:
        # Fallback: if signed URL fails, indicate client should use /pdf endpoint
        print(f"  [WARN] Signed URL failed for {notice_id}: {e}")
        raise HTTPException(500, f"Signed URL failed: {str(e)[:100]}")


# ── AI Q&A ──────────────────────────────────────────────

class QaRequest(BaseModel):
    notice_id: str
    messages: list  # [{role:'user'|'ai', text:''}]
    question: str

@app.post("/api/qa")
async def qa_chat(req: QaRequest, request: Request):
    user = await get_current_user(request)
    await check_notice_access(req.notice_id, user)
    r = db_get("notices", req.notice_id)
    if not r: raise HTTPException(404)
    header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
    items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")
    real_items = [it for it in items if not it.get("is_subtotal")]

    ctx = f"""You are an expert LP fund operations assistant. Answer about this notice.
Notice: {header.get('notice_title','')}
Fund: {header.get('Underlying_Fund_Name_full','')}
LP: {header.get('LP_Name_full','')}
Date: {header.get('issue_date','')} Due: {header.get('due_date','')}
Net Amount: {header.get('LP_net_amount')}
Items: {json.dumps([{'name':it.get('item_name'),'amount':it.get('LP_signed_amount'),'type':it.get('Transaction_type'),'commit':it.get('Commitment_affecting')} for it in real_items], ensure_ascii=False)}
Header: {json.dumps(header, ensure_ascii=False)}
Answer in the user's language. Be concise."""

    messages = [
        {"role": "user", "parts": [{"text": ctx}]},
        {"role": "model", "parts": [{"text": "네, 이 Notice에 대해 질문해주세요."}]}
    ]
    for m in req.messages:
        messages.append({"role": "user" if m["role"]=="user" else "model", "parts": [{"text": m["text"]}]})
    messages.append({"role": "user", "parts": [{"text": req.question}]})

    oid = user.get("org_id", "")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{_get_setting('gemini_model', GEMINI_MODEL, org_id=oid)}:generateContent?key={GEMINI_KEY}"
    qa_temp = float(_get_setting("gemini_temperature", str(GEMINI_TEMPERATURE), org_id=oid)) + 0.2  # slightly higher for chat
    qa_temp = min(qa_temp, 2.0)
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(url, json={
            "contents": messages,
            "generationConfig": {"temperature": qa_temp, "maxOutputTokens": 2048}
        })
    if resp.status_code != 200:
        raise HTTPException(502, f"Gemini error: {resp.text[:200]}")
    data = resp.json()
    ai_text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "응답 생성 실패")
    return {"answer": ai_text}


# ── Models list ─────────────────────────────────────────
@app.get("/api/models")
async def list_models(request: Request):
    user = await get_current_user(request)
    url = f"https://generativelanguage.googleapis.com/v1beta/models?key={GEMINI_KEY}"
    async with httpx.AsyncClient(timeout=15.0) as client:
        resp = await client.get(url)
    if resp.status_code != 200:
        raise HTTPException(502, "Failed to fetch models")
    data = resp.json()
    models = [{"id": m["name"].replace("models/",""), "displayName": m.get("displayName",""),
               "outLimit": m.get("outputTokenLimit", 8192)}
              for m in data.get("models",[])
              if "generateContent" in (m.get("supportedGenerationMethods") or [])]
    return models


# ── Settings ────────────────────────────────────────────

@app.get("/api/settings")
async def get_settings(request: Request):
    user = await get_current_user(request)
    """Return all AI-related settings (per-org)."""
    oid = user.get("org_id", "")
    return {
        "gemini_model": _get_setting("gemini_model", GEMINI_MODEL, org_id=oid),
        "gemini_temperature": float(_get_setting("gemini_temperature", str(GEMINI_TEMPERATURE), org_id=oid)),
    }

@app.put("/api/settings")
async def update_settings(body: dict, request: Request):
    user = await get_current_user(request)
    require_uploader(user)
    """Update AI-related settings (per-org)."""
    oid = user.get("org_id", "")
    allowed = {"gemini_model", "gemini_temperature"}
    for k, v in body.items():
        if k in allowed:
            _set_setting(k, str(v), org_id=oid)
    return {
        "gemini_model": _get_setting("gemini_model", GEMINI_MODEL, org_id=oid),
        "gemini_temperature": float(_get_setting("gemini_temperature", str(GEMINI_TEMPERATURE), org_id=oid)),
    }


# ── Asset Groups (AI-based grouping) ──────────────────

ASSET_GROUP_PROMPT_PASS1 = """You are a private equity fund operations expert. Given a list of asset/investment names extracted from Capital Call and Distribution notices, group them by the SAME underlying investment/company/asset.

CRITICAL RULES — be AGGRESSIVE in grouping:
1. The SAME company/asset appears with many different prefixes and suffixes. ALL of these refer to the SAME asset and MUST be grouped:
   - Transaction type prefixes: "Capital Call - X", "Distribution - X", "Current Call - X", "Reallocation Call - X", "Reallocation Distribution - X", "Current Distribution - X"
   - Parenthetical suffixes: "X (Reallocation)", "X (Prior Distributions)", "X (Offset)", "X (Distribution)"
   - Colon prefixes: "Current Capital Call: X", "Reallocation of Prior Capital Calls: X"
   - Combined: "Reallocation - X (Prior Distributions)" and "Current Call - X (Offset)" both refer to X

2. CONCRETE EXAMPLES of items that MUST be in ONE group:
   "Capital Call - Yondr", "Reallocation Call - Yondr", "Current Distribution - Yondr", 
   "Reallocation Distribution - Yondr", "Distribution - Yondr - Ticking Fee"
   → ALL are ONE group: "Yondr"
   
   "Capital Call - Management Fees", "Reallocation Call - Management Fees"
   → ONE group: "Management Fees"
   
   "Reallocation - Vantage NA DevCo", "Current Call - Vantage NA DevCo", 
   "Distribution - Vantage NA DevCo", "Reallocation Distribution - Vantage NA DevCo",
   "Reallocation - Vantage NA DevCo (Prior Distributions)"
   → ALL ONE group: "Vantage NA DevCo"

3. Strip ALL transaction-type prefixes to find the underlying asset name. The core asset name is what comes AFTER the prefix.

4. Fee/expense items with the SAME base name but different prefixes are ONE group:
   "Capital Call - Partnership Expenses" + "Reallocation Call - Partnership Expenses" → "Partnership Expenses"
   "Capital Call - Organizational Expenses" + "Current Call - Excess Organizational Expenses" → "Organizational Expenses"

5. Choose the SHORTEST clean name (without prefix/suffix) as the canonical group key.

6. Items that genuinely have no related variants remain as singletons.

Return ONLY valid JSON (no markdown/backticks):
{"groups": {"Canonical Name": ["raw name 1", "raw name 2"], "Another": ["raw name 3"]}}

Every input name must appear in exactly one group. Do not omit any.

Input asset names:
"""

ASSET_GROUP_PROMPT_PASS2 = """You previously grouped asset names into the groups shown below. However, some items remained as singletons (1 member each). 

Review the singletons and check if any of them should be merged into an existing group. The underlying asset name may be hidden behind a transaction-type prefix like "Capital Call - ", "Distribution - ", "Reallocation - ", "Current Call - ", etc.

IMPORTANT: Be aggressive — if a singleton's core asset name matches an existing group, merge it.

Current groups (canonical → members):
{existing_groups}

Singleton items to review:
{singletons}

Return ONLY valid JSON with the COMPLETE updated grouping (all groups, including unchanged ones):
{"groups": {"Canonical": ["member1", "member2", ...], ...}}

Every name from both the existing groups and singletons must appear exactly once. Do not omit any.
"""

ASSET_GROUP_TEMPERATURE = 0.4  # Higher than parsing (0.1) for more aggressive grouping


def _parse_gemini_json(raw_text: str) -> dict:
    """Extract JSON from Gemini response text."""
    js = raw_text
    mt_idx = raw_text.find("```")
    if mt_idx >= 0:
        end_idx = raw_text.find("```", mt_idx + 3)
        if end_idx > mt_idx:
            js = raw_text[mt_idx+3:end_idx]
            if js.startswith("json"):
                js = js[4:]
    bs, be = js.find("{"), js.rfind("}")
    if bs == -1 or be == -1:
        raise ValueError("No JSON found in response")
    return json.loads(js[bs:be+1])


async def _call_gemini_text(prompt: str, model: str = None, temperature: float = None, org_id: str = None) -> str:
    """Call Gemini with text-only prompt, return raw text response."""
    _model = model or _get_setting("gemini_model", GEMINI_MODEL, org_id=org_id)
    _temp = temperature if temperature is not None else ASSET_GROUP_TEMPERATURE
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{_model}:generateContent?key={GEMINI_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": _temp, "maxOutputTokens": 16384}
    }
    async with httpx.AsyncClient(timeout=90.0) as client:
        resp = await client.post(url, json=payload)
    if resp.status_code != 200:
        raise Exception(f"Gemini API error: {resp.status_code} {resp.text[:200]}")
    data = resp.json()
    return data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")


@app.post("/api/asset-groups")
async def create_asset_groups(body: dict, request: Request):
    user = await get_current_user(request)
    require_uploader(user)
    """AI-based asset name grouping with two-pass strategy. Returns SSE stream."""
    fund_key = body.get("fund_key", "")
    asset_names = body.get("asset_names", [])
    model = body.get("model")

    if not asset_names:
        raise HTTPException(400, "No asset names provided")

    async def sse_stream():
        try:
            yield f"data: {json.dumps({'stage': 'progress', 'pct': 0, 'message': 'AI 그룹화 시작 (Pass 1/2)...'})}\n\n"

            # ── Pass 1: Initial grouping ──
            names_text = "\n".join(f"- {name}" for name in asset_names)
            prompt1 = ASSET_GROUP_PROMPT_PASS1 + names_text

            # Call with heartbeat
            gemini_task = asyncio.create_task(_call_gemini_text(prompt1, model, org_id=user["org_id"]))
            hb_count = 0
            while not gemini_task.done():
                try:
                    await asyncio.wait_for(asyncio.shield(gemini_task), timeout=5)
                except asyncio.TimeoutError:
                    hb_count += 1
                    pct = min(5 + hb_count * 4, 45)
                    yield f"data: {json.dumps({'stage': 'progress', 'pct': pct, 'message': f'Pass 1 AI 분석 중... ({hb_count * 5}s)'})}\n\n"
                except Exception:
                    break

            raw_text1 = gemini_task.result()
            parsed1 = _parse_gemini_json(raw_text1)
            groups = parsed1.get("groups", parsed1)

            # Validate coverage
            covered = set()
            for canonical, members in groups.items():
                if isinstance(members, list):
                    for m in members:
                        covered.add(m)
            for m in asset_names:
                if m not in covered:
                    groups[m] = [m]

            # Count singletons
            singletons = {k: v for k, v in groups.items() if isinstance(v, list) and len(v) == 1}
            multi_groups = {k: v for k, v in groups.items() if isinstance(v, list) and len(v) > 1}

            yield f"data: {json.dumps({'stage': 'progress', 'pct': 50, 'message': f'Pass 1 완료: {len(multi_groups)} 그룹, {len(singletons)} singleton'})}\n\n"

            # ── Pass 2: Merge remaining singletons ──
            if len(singletons) >= 2 and len(multi_groups) >= 1:
                yield f"data: {json.dumps({'stage': 'progress', 'pct': 55, 'message': f'Pass 2 시작: {len(singletons)}개 singleton 재검토...'})}\n\n"

                existing_desc = "\n".join(
                    f'  "{k}": {json.dumps(v)}' for k, v in multi_groups.items()
                )
                singleton_list = "\n".join(f"- {list(v)[0]}" for v in singletons.values())

                prompt2 = ASSET_GROUP_PROMPT_PASS2.replace("{existing_groups}", existing_desc).replace("{singletons}", singleton_list)

                gemini_task2 = asyncio.create_task(_call_gemini_text(prompt2, model, org_id=user["org_id"]))
                hb_count2 = 0
                while not gemini_task2.done():
                    try:
                        await asyncio.wait_for(asyncio.shield(gemini_task2), timeout=5)
                    except asyncio.TimeoutError:
                        hb_count2 += 1
                        pct = min(55 + hb_count2 * 4, 90)
                        yield f"data: {json.dumps({'stage': 'progress', 'pct': pct, 'message': f'Pass 2 AI 분석 중... ({hb_count2 * 5}s)'})}\n\n"
                    except Exception:
                        break

                try:
                    raw_text2 = gemini_task2.result()
                    parsed2 = _parse_gemini_json(raw_text2)
                    groups2 = parsed2.get("groups", parsed2)

                    # Validate pass 2 didn't lose any names
                    covered2 = set()
                    for members in groups2.values():
                        if isinstance(members, list):
                            for m in members:
                                covered2.add(m)
                    all_names = set(asset_names)
                    if covered2 >= all_names:
                        # Pass 2 covers everything — use it
                        groups = groups2
                        singletons2 = {k: v for k, v in groups.items() if isinstance(v, list) and len(v) == 1}
                        multi_groups2 = {k: v for k, v in groups.items() if isinstance(v, list) and len(v) > 1}
                        yield f"data: {json.dumps({'stage': 'progress', 'pct': 92, 'message': f'Pass 2 완료: {len(multi_groups2)} 그룹, {len(singletons2)} singleton'})}\n\n"
                    else:
                        # Pass 2 lost some names — keep pass 1 result
                        missing2 = all_names - covered2
                        print(f"  [WARN] Pass 2 lost {len(missing2)} names, keeping pass 1 result")
                        yield f"data: {json.dumps({'stage': 'progress', 'pct': 92, 'message': 'Pass 2 불완전 — Pass 1 결과 유지'})}\n\n"
                except Exception as e2:
                    print(f"  [WARN] Pass 2 failed: {e2}, keeping pass 1 result")
                    yield f"data: {json.dumps({'stage': 'progress', 'pct': 92, 'message': f'Pass 2 실패 — Pass 1 결과 유지'})}\n\n"
            else:
                yield f"data: {json.dumps({'stage': 'progress', 'pct': 92, 'message': 'Pass 2 불필요 (singleton 부족)'})}\n\n"

            # ── Save (per-user) ──
            yield f"data: {json.dumps({'stage': 'progress', 'pct': 95, 'message': '저장 중...'})}\n\n"

            uid = user["id"]
            oid = user.get("org_id", "")
            existing_ag = get_supa().table("asset_groups").select("fund_key").eq("fund_key", fund_key).eq("org_id", oid).limit(1).execute()
            if existing_ag.data:
                get_supa().table("asset_groups").update({"groups_json": groups}).eq("fund_key", fund_key).eq("org_id", oid).execute()
            else:
                get_supa().table("asset_groups").insert({"fund_key": fund_key, "groups_json": groups, "org_id": oid}).execute()

            final_singletons = sum(1 for v in groups.values() if isinstance(v, list) and len(v) == 1)
            final_merged = sum(1 for v in groups.values() if isinstance(v, list) and len(v) > 1)
            yield f"data: {json.dumps({'stage': 'done', 'groups': groups, 'total': len(asset_names), 'group_count': len(groups), 'merged_count': final_merged, 'singleton_count': final_singletons})}\n\n"

        except json.JSONDecodeError as e:
            yield f"data: {json.dumps({'stage': 'error', 'detail': f'JSON 파싱 오류: {str(e)}'})}\n\n"
        except Exception as e:
            import traceback; traceback.print_exc()
            yield f"data: {json.dumps({'stage': 'error', 'detail': str(e)[:500]})}\n\n"

    return StreamingResponse(sse_stream(), media_type="text/event-stream")


@app.get("/api/asset-groups/{fund_key}")
async def get_asset_groups(fund_key: str, request: Request):
    user = await get_current_user(request)
    """Retrieve saved asset groups for a fund (per-org)."""
    oid = user.get("org_id", "")
    r = get_supa().table("asset_groups").select("*").eq("fund_key", fund_key).eq("org_id", oid).limit(1).execute()
    row = r.data[0] if r.data else None
    if not row:
        return {"fund_key": fund_key, "groups": None}
    return {
        "fund_key": fund_key,
        "groups": row["groups_json"] if isinstance(row["groups_json"], dict) else json.loads(row["groups_json"]),
        "updated_at": row["updated_at"],
    }


@app.put("/api/asset-groups/{fund_key}")
async def save_asset_groups(fund_key: str, body: dict, request: Request):
    user = await get_current_user(request)
    require_uploader(user)
    """Save/update asset groups for a fund (per-org)."""
    oid = user.get("org_id", "")
    groups = body.get("groups", {})
    existing = get_supa().table("asset_groups").select("fund_key").eq("fund_key", fund_key).eq("org_id", oid).limit(1).execute()
    if existing.data:
        get_supa().table("asset_groups").update({"groups_json": groups}).eq("fund_key", fund_key).eq("org_id", oid).execute()
    else:
        get_supa().table("asset_groups").insert({"fund_key": fund_key, "groups_json": groups, "org_id": oid}).execute()
    return {"ok": True, "fund_key": fund_key}


# ── Auth Info ──────────────────────────────────────────
@app.get("/api/auth/me")
async def auth_me(request: Request):
    """Return current user info + role + org + name."""
    user = await get_current_user(request)
    org = None
    if user.get("org_id"):
        org_row = db_get("organizations", user["org_id"])
        if org_row:
            org = {"id": org_row["id"], "name": org_row["name"]}
    # Fetch name fields
    role_row = db_get("user_roles", user["id"], id_col="user_id")
    name_en = f"{role_row.get('name_en_first','')} {role_row.get('name_en_last','')}".strip() if role_row else ""
    name_kr = f"{role_row.get('name_kr_last','')}{role_row.get('name_kr_first','')}".strip() if role_row else ""
    return {"id": user["id"], "email": user["email"], "role": user["role"],
            "org_id": user.get("org_id"), "org": org,
            "name_en": name_en, "name_kr": name_kr}


# ── Organizations API ─────────────────────────────────

@app.get("/api/organizations")
async def list_organizations():
    """List organizations available for registration (is_system=FALSE only). No auth required."""
    rows = db_list("organizations", order_col="created_at", order_desc=False)
    return [{"id": r["id"], "name": r["name"]} for r in rows if not r.get("is_system")]


@app.post("/api/auth/register-org")
async def register_org(body: dict, request: Request):
    """Assign org + name to user after signup. Creates new org if org_name provided."""
    user = await get_current_user(request)
    org_id = body.get("org_id")
    org_name = body.get("org_name", "").strip()

    # Name fields
    name_en_last = body.get("name_en_last", "").strip()
    name_en_first = body.get("name_en_first", "").strip()
    name_kr_last = body.get("name_kr_last", "").strip()
    name_kr_first = body.get("name_kr_first", "").strip()

    if not org_id and not org_name:
        raise HTTPException(400, "org_id or org_name required")

    if org_name:
        # Create new organization
        new_id = "org_" + org_name.lower().replace(" ", "_").replace("-", "_")[:30]
        existing = db_get("organizations", new_id)
        if existing:
            org_id = new_id  # Already exists, just join
        else:
            db_insert("organizations", {"id": new_id, "name": org_name, "is_system": False})
            org_id = new_id
            print(f"  [ORG] New organization created: {org_name} ({new_id})")

    # Update user's org_id + name
    update_data = {"org_id": org_id}
    if name_en_last: update_data["name_en_last"] = name_en_last
    if name_en_first: update_data["name_en_first"] = name_en_first
    if name_kr_last: update_data["name_kr_last"] = name_kr_last
    if name_kr_first: update_data["name_kr_first"] = name_kr_first
    db_update("user_roles", update_data, user["id"], id_col="user_id")
    return {"ok": True, "org_id": org_id}


# ── Favorites API ─────────────────────────────────────

@app.get("/api/favorites")
async def get_favorites(request: Request):
    """Get user's subscribed fund keys."""
    user = await get_current_user(request)
    rows = db_list("user_fund_favorites", user_id=user["id"])
    return [r["fund_key"] for r in rows]


@app.put("/api/favorites")
async def save_favorites(body: dict, request: Request):
    """Save user's subscribed fund keys (full replace)."""
    user = await get_current_user(request)
    fund_keys = body.get("fund_keys", [])
    # Delete all existing
    get_supa().table("user_fund_favorites").delete().eq("user_id", user["id"]).execute()
    # Insert new
    for fk in fund_keys:
        if fk and isinstance(fk, str):
            db_insert("user_fund_favorites", {"user_id": user["id"], "fund_key": fk})
    return {"ok": True, "fund_keys": fund_keys}


# ── Admin API ──────────────────────────────────────────

@app.get("/api/admin/organizations")
async def admin_list_organizations(request: Request):
    """List all organizations with stats. Admin only.
    notice_count uses unique duplicate_key to avoid counting copies."""
    admin = await require_admin(request)
    orgs = db_list("organizations", order_col="created_at", order_desc=False)
    result = []
    for o in orgs:
        notices = db_list("notices", org_id=o["id"])
        # Count unique notices (by duplicate_key, fallback to id)
        seen_keys = set()
        for n in notices:
            key = n.get("duplicate_key") or n.get("pdf_hash") or n["id"]
            seen_keys.add(key)
        user_count = len(db_list("user_roles", org_id=o["id"]))
        result.append({
            "id": o["id"], "name": o["name"], "is_system": o.get("is_system", False),
            "notice_count": len(seen_keys), "user_count": user_count,
        })
    # Calculate total unique notices across all orgs
    all_notices = db_list("notices")
    all_keys = set()
    for n in all_notices:
        key = n.get("duplicate_key") or n.get("pdf_hash") or n["id"]
        all_keys.add(key)
    return {"orgs": result, "total_unique": len(all_keys)}

@app.get("/api/admin/users")
async def admin_list_users(request: Request):
    """List all users with notice counts. Admin only."""
    admin = await require_admin(request)
    roles = db_list("user_roles", order_col="created_at", order_desc=False)
    result = []
    for r in roles:
        # Count notices for this user's org
        org_id = r.get("org_id", "")
        org_row = db_get("organizations", org_id) if org_id else None
        result.append({
            "user_id": r["user_id"],
            "email": r.get("email", ""),
            "role": r.get("role", "uploader"),
            "org_id": org_id,
            "org_name": org_row["name"] if org_row else "",
            "name_en": f"{r.get('name_en_first','')} {r.get('name_en_last','')}".strip(),
            "name_kr": f"{r.get('name_kr_last','')}{r.get('name_kr_first','')}".strip(),
            "created_at": r.get("created_at"),
        })
    return result

@app.get("/api/admin/users/{uid}/notices")
async def admin_user_notices(uid: str, request: Request):
    """List notices for a specific user's org. Admin only."""
    admin = await require_admin(request)
    # Get user's org
    ur = db_get("user_roles", uid, id_col="user_id")
    org_id = ur.get("org_id", "") if ur else ""
    if org_id:
        rows = db_list("notices", order_col="created_at", order_desc=True, org_id=org_id)
    else:
        rows = db_list("notices", order_col="created_at", order_desc=True, user_id=uid)
    result = []
    for r in rows:
        header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
        line_items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")
        result.append({
            "id": r["id"], "fileName": r.get("file_name",""), "analyzedAt": r.get("analyzed_at"),
            "header": header, "lineItems": line_items,
            "is_voided": bool(r.get("is_voided")), "voided_by": r.get("voided_by"),
            "page_count": r.get("page_count") or 0,
        })
    return result

@app.put("/api/admin/users/{uid}/role")
async def admin_update_role(uid: str, body: dict, request: Request):
    """Change user role (uploader ↔ viewer only). Admin only."""
    admin = await require_admin(request)
    new_role = body.get("role", "uploader")
    if new_role not in ("uploader", "viewer"):
        raise HTTPException(400, "Role must be 'uploader' or 'viewer'")
    # Cannot change another admin's role
    target = db_get("user_roles", uid, id_col="user_id")
    if target and target.get("role") == "admin":
        raise HTTPException(400, "Cannot change admin role from frontend")
    db_update("user_roles", {"role": new_role}, uid, id_col="user_id")
    return {"ok": True, "user_id": uid, "role": new_role}

@app.delete("/api/admin/users/{uid}")
async def admin_delete_user(uid: str, request: Request):
    """Delete a user and their favorites. Admin only. Notices belong to org, not deleted."""
    admin = await require_admin(request)
    if uid == admin["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    target = db_get("user_roles", uid, id_col="user_id")
    if target and target.get("role") == "admin":
        raise HTTPException(400, "Cannot delete another admin")
    # Delete favorites
    try: get_supa().table("user_fund_favorites").delete().eq("user_id", uid).execute()
    except: pass
    # Delete role entry
    try: db_delete("user_roles", uid, id_col="user_id")
    except: pass
    return {"ok": True}


# ── Admin Notice Assignment ───────────────────────────

@app.put("/api/admin/notices/{notice_id}/move")
async def admin_move_notice(notice_id: str, body: dict, request: Request):
    """Move a notice to a different organization. Admin only."""
    admin = await require_admin(request)
    target_org = body.get("org_id", "")
    if not target_org:
        raise HTTPException(400, "org_id required")
    # Verify notice exists
    r = db_get("notices", notice_id)
    if not r:
        raise HTTPException(404, "Notice not found")
    # Verify target org exists
    org = db_get("organizations", target_org)
    if not org:
        raise HTTPException(404, "Organization not found")
    # Update org_id
    db_update("notices", {"org_id": target_org}, notice_id)
    print(f"  [ADMIN] Moved notice {notice_id} → {target_org}")
    return {"ok": True, "notice_id": notice_id, "org_id": target_org}


@app.post("/api/admin/notices/{notice_id}/copy")
async def admin_copy_notice(notice_id: str, body: dict, request: Request):
    """Copy a notice to one or more organizations. Admin only.
    Creates independent copies (separate DB records + storage files)."""
    admin = await require_admin(request)
    target_orgs = body.get("org_ids", [])
    if not target_orgs:
        raise HTTPException(400, "org_ids required (list)")

    # Verify source notice
    r = db_get("notices", notice_id)
    if not r:
        raise HTTPException(404, "Notice not found")

    header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
    line_items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")

    created = []
    for org_id in target_orgs:
        org = db_get("organizations", org_id)
        if not org or org.get("is_system"):
            continue
        # Skip if same org as source
        if org_id == r.get("org_id"):
            continue
        # Check if duplicate already exists in target org
        if r.get("duplicate_key"):
            existing = get_supa().table("notices").select("id").eq("duplicate_key", r["duplicate_key"]).eq("org_id", org_id).limit(1).execute()
            if existing.data:
                print(f"  [ADMIN] Skip copy to {org_id}: duplicate exists ({existing.data[0]['id']})")
                continue

        # Generate new notice ID
        new_id = f"{notice_id}_cp{org_id[-8:]}"
        # Remove if somehow exists
        try: db_delete("notices", new_id)
        except: pass

        # Copy DB record
        db_insert("notices", {
            "id": new_id,
            "file_name": r.get("file_name", ""),
            "user_id": admin["id"],
            "org_id": org_id,
            "header": header,
            "line_items": line_items,
            "raw_ai_response": r.get("raw_ai_response", ""),
            "pdf_hash": r.get("pdf_hash", ""),
            "page_count": r.get("page_count", 0),
            "duplicate_key": r.get("duplicate_key", ""),
        })

        # Copy storage files (PDF + text_map)
        copy_storage(notice_id, new_id)

        created.append({"notice_id": new_id, "org_id": org_id, "org_name": org["name"]})
        print(f"  [ADMIN] Copied notice {notice_id} → {new_id} (org: {org_id})")

    return {"ok": True, "source_id": notice_id, "created": created}


@app.get("/api/admin/notice-org-map")
async def admin_notice_org_map(request: Request):
    """Return all notices with org assignment info for admin portfolio management.
    Groups notices by pdf_hash to identify same-PDF across orgs."""
    admin = await require_admin(request)
    all_notices = db_list("notices", order_col="created_at", order_desc=True)
    orgs = [o for o in db_list("organizations", order_col="created_at") if not o.get("is_system")]

    # Build map: pdf_hash -> list of (notice_id, org_id)
    result = []
    for r in all_notices:
        header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
        result.append({
            "id": r["id"],
            "file_name": r.get("file_name", ""),
            "org_id": r.get("org_id", ""),
            "pdf_hash": r.get("pdf_hash", ""),
            "duplicate_key": r.get("duplicate_key", ""),
            "fund": header.get("Underlying_Fund_Name_short", "") or header.get("Fund_ID_Key", ""),
            "fund_key": header.get("Fund_ID_Key", ""),
            "lp_short": header.get("LP_Name_short", ""),
            "issue_date": header.get("issue_date", ""),
            "notice_type": header.get("notice_type", ""),
            "net_amount": header.get("LP_net_amount"),
        })

    return {"notices": result, "orgs": [{"id": o["id"], "name": o["name"]} for o in orgs]}


@app.post("/api/admin/notice-org-toggle")
async def admin_notice_org_toggle(body: dict, request: Request):
    """Toggle a notice's presence in a specific org.
    If notice exists in org -> remove it. If not -> copy it there."""
    admin = await require_admin(request)
    notice_id = body.get("notice_id", "")
    target_org = body.get("org_id", "")
    action = body.get("action", "")  # 'add' or 'remove'

    if not notice_id or not target_org or action not in ("add", "remove"):
        raise HTTPException(400, "notice_id, org_id, and action ('add'/'remove') required")

    if action == "remove":
        # Find the notice copy in this org and delete it
        r = db_get("notices", notice_id)
        if r and r.get("org_id") == target_org:
            db_delete("notices", notice_id)
            delete_storage(notice_id)
            print(f"  [ADMIN] Removed notice {notice_id} from {target_org}")
            return {"ok": True, "action": "removed", "notice_id": notice_id}
        # Maybe it's a different copy - find by pdf_hash in that org
        return {"ok": False, "detail": "Notice not found in that org"}

    elif action == "add":
        # Copy the source notice to target org
        source = db_get("notices", notice_id)
        if not source:
            raise HTTPException(404, "Source notice not found")

        org = db_get("organizations", target_org)
        if not org:
            raise HTTPException(404, "Organization not found")

        # Check if already exists in target org (by duplicate_key)
        if source.get("duplicate_key"):
            existing = get_supa().table("notices").select("id").eq("duplicate_key", source["duplicate_key"]).eq("org_id", target_org).limit(1).execute()
            if existing.data:
                return {"ok": True, "action": "already_exists", "existing_id": existing.data[0]["id"]}

        header = source["header"] if isinstance(source["header"], dict) else json.loads(source["header"] or "{}")
        line_items = source["line_items"] if isinstance(source["line_items"], list) else json.loads(source["line_items"] or "[]")

        new_id = f"{notice_id}_cp{target_org[-8:]}"
        try: db_delete("notices", new_id)
        except: pass

        db_insert("notices", {
            "id": new_id,
            "file_name": source.get("file_name", ""),
            "user_id": admin["id"],
            "org_id": target_org,
            "header": header,
            "line_items": line_items,
            "raw_ai_response": source.get("raw_ai_response", ""),
            "pdf_hash": source.get("pdf_hash", ""),
            "page_count": source.get("page_count", 0),
            "duplicate_key": source.get("duplicate_key", ""),
        })
        copy_storage(notice_id, new_id)

        print(f"  [ADMIN] Added notice copy {new_id} to {target_org}")
        return {"ok": True, "action": "added", "new_id": new_id, "org_id": target_org}


# ── Health ──────────────────────────────────────────────
@app.get("/api/health")
async def health():
    return {"status": "ok", "notices": _count_notices()}

def _count_notices():
    return db_count("notices")


# ── Excel Export (openpyxl) ────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, date as dt_date

# Style constants
_XS_NAVY = 'FF1D3557'
_XS_NAVY_LIGHT = 'FFE8EDF3'
_XS_GREEN = 'FF4E9A6D'
_XS_AMBER = 'FFD97706'
_XS_GRAY = 'FFF5F6F8'
_XS_BDR = 'FFE2E8F0'
_XS_WHITE = 'FFFFFFFF'

_FNT_TITLE = Font(name='Calibri', size=14, bold=True, color=_XS_WHITE)
_FNT_SEC = Font(name='Calibri', size=10, bold=True, color=_XS_NAVY)
_FNT_HDR = Font(name='Calibri', size=9, bold=True, color='FF6B7280')
_FNT_BOLD = Font(name='Calibri', size=10, bold=True)
_FNT_NORM = Font(name='Calibri', size=10)
_FNT_SMALL = Font(name='Calibri', size=9, color='FF6B7280')

_FILL_NAVY = PatternFill('solid', fgColor=_XS_NAVY)
_FILL_LIGHT = PatternFill('solid', fgColor=_XS_NAVY_LIGHT)
_FILL_GRAY = PatternFill('solid', fgColor=_XS_GRAY)
_FILL_CUR = PatternFill('solid', fgColor='FFDBEAFE')  # 현재 notice 하이라이트

_ALIGN_R = Alignment(horizontal='right', vertical='center')
_ALIGN_L = Alignment(horizontal='left', vertical='center')
_ALIGN_C = Alignment(horizontal='center', vertical='center')

_BDR_THIN = Border(bottom=Side('thin', color=_XS_BDR))
_BDR_MED = Border(bottom=Side('medium', color=_XS_NAVY))
_BDR_TOP_DBL = Border(top=Side('double', color='FF000000'))

_NUM = '#,##0.00'
_NUM_ACCT = '#,##0.00_);(#,##0.00)'
_PCT = '0.0%'
_DATE = 'YYYY-MM-DD'


def _xs_date(s):
    """Parse date string to datetime.date for Excel."""
    if not s or s == 'N/A':
        return None
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except:
        return s


def _xs_set_row(ws, row, data, styles=None):
    """Write a row of data with optional per-cell styles."""
    for ci, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        if styles:
            s = styles[ci - 1] if ci - 1 < len(styles) else None
            if s:
                if 'font' in s: cell.font = s['font']
                if 'fill' in s: cell.fill = s['fill']
                if 'align' in s: cell.alignment = s['align']
                if 'fmt' in s: cell.number_format = s['fmt']
                if 'border' in s: cell.border = s['border']


def _xs_style_header_row(ws, row, max_col):
    """Apply header style to a row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _FNT_HDR
        cell.fill = _FILL_LIGHT
        cell.border = _BDR_MED
        if c >= 4:  # numeric columns
            cell.alignment = _ALIGN_R


def _xs_notice_sheet(wb, notice, sheet_name='Notice'):
    """Build a single Notice sheet with formatting and formulas."""
    ws = wb.create_sheet(sheet_name)
    h = notice['header']
    items = [it for it in notice.get('line_items', []) if not it.get('is_subtotal')]

    # Title row (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1,
        value=f"{h.get('Underlying_Fund_Name_full', '')} — {h.get('notice_type', '')} Notice")
    title_cell.font = _FNT_TITLE
    title_cell.fill = _FILL_NAVY
    title_cell.alignment = _ALIGN_L
    for c in range(2, 6):
        ws.cell(row=1, column=c).fill = _FILL_NAVY
    ws.row_dimensions[1].height = 30

    # Header info
    r = 3
    info_fields = [
        ('Fund Name', h.get('Underlying_Fund_Name_full', '')),
        ('Notice Title', h.get('notice_title', '')),
        ('Notice Number', h.get('notice_number', '')),
        ('Issue Date', _xs_date(h.get('issue_date', ''))),
        ('Due Date', _xs_date(h.get('due_date', ''))),
        ('LP Name', h.get('LP_Name_full', '')),
        ('LP Code', h.get('LP_code', '')),
        ('Commitment', _pn(h.get('Commitment_original'))),
        ('LP Interest %', (_pn(h.get('pct_LP_Interest')) or 0) / 100 if _pn(h.get('pct_LP_Interest')) else None),
    ]
    for label, val in info_fields:
        ws.cell(row=r, column=1, value=label).font = _FNT_BOLD
        ws.cell(row=r, column=1).fill = _FILL_GRAY
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = _FNT_NORM
        if isinstance(val, (int, float)) and val and label == 'Commitment':
            cell.number_format = _NUM
        elif label == 'LP Interest %' and val is not None:
            cell.number_format = _PCT
        elif isinstance(val, dt_date):
            cell.number_format = _DATE
        r += 1

    # Section: Prior
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value='Prior to Current Notice').font = _FNT_SEC
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = _FILL_LIGHT
    r += 1
    for label, key in [('Unfunded Commitment', 'Unfunded_prior'), ('Cumul. Contributions', 'CumContribPrior'), ('Cumul. Distributions', 'CumDistribPrior')]:
        ws.cell(row=r, column=1, value=label).font = _FNT_BOLD
        ws.cell(row=r, column=1).fill = _FILL_GRAY
        cell = ws.cell(row=r, column=2, value=_pn(h.get(key)))
        cell.number_format = _NUM_ACCT
        r += 1

    # Section: Current
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value='Current Notice').font = _FNT_SEC
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = _FILL_LIGHT
    r += 1
    for label, key in [('Current Contributions', 'Current_Commit_Contribution'), ('Current Distributions', 'Current_Commit_Distribution'), ('LP Net Amount', 'LP_net_amount')]:
        ws.cell(row=r, column=1, value=label).font = _FNT_BOLD
        ws.cell(row=r, column=1).fill = _FILL_GRAY
        cell = ws.cell(row=r, column=2, value=_pn(h.get(key)))
        cell.number_format = _NUM_ACCT
        if label == 'LP Net Amount':
            cell.font = _FNT_BOLD
        r += 1

    # Section: After
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value='After Current Notice').font = _FNT_SEC
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = _FILL_LIGHT
    r += 1
    for label, key in [('Unfunded Commitment', 'Unfunded_after'), ('Cumul. Contributions', 'CumContribAfter'), ('Cumul. Distributions', 'CumDistribAfter')]:
        ws.cell(row=r, column=1, value=label).font = _FNT_BOLD
        ws.cell(row=r, column=1).fill = _FILL_GRAY
        cell = ws.cell(row=r, column=2, value=_pn(h.get(key)))
        cell.number_format = _NUM_ACCT
        r += 1

    # Line Items
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value='Line Items').font = _FNT_SEC
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = _FILL_LIGHT
    r += 1
    hdr_row = r
    for ci, txt in enumerate(['#', 'Item Name', 'LP Amount', 'Type', 'Commit'], 1):
        c = ws.cell(row=r, column=ci, value=txt)
        c.font = _FNT_HDR
        c.fill = _FILL_LIGHT
        c.border = _BDR_MED
        if ci == 3: c.alignment = _ALIGN_R
        elif ci >= 4: c.alignment = _ALIGN_C
    r += 1

    item_start = r
    for idx, it in enumerate(items):
        ws.cell(row=r, column=1, value=idx + 1).font = _FNT_NORM
        ws.cell(row=r, column=2, value=it.get('item_name', '')).font = _FNT_NORM
        amt_cell = ws.cell(row=r, column=3, value=_pn(it.get('LP_signed_amount')))
        amt_cell.number_format = _NUM_ACCT
        amt_cell.alignment = _ALIGN_R
        ws.cell(row=r, column=4, value='Call' if it.get('Transaction_type') == 'call' else 'Dist').alignment = _ALIGN_C
        ws.cell(row=r, column=5, value='Yes' if it.get('Commitment_affecting') else 'No').alignment = _ALIGN_C
        # Alternate row shading
        if idx % 2 == 1:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = _FILL_GRAY
        r += 1
    item_end = r - 1

    # Summary with FORMULAS
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value='Summary').font = _FNT_SEC
    ws.cell(row=r, column=1).fill = _FILL_LIGHT
    for c in range(2, 6):
        ws.cell(row=r, column=c).fill = _FILL_LIGHT
    r += 1

    col_amt = get_column_letter(3)
    col_type = get_column_letter(4)
    col_ca = get_column_letter(5)

    summary_items = [
        ('Commitment Affecting — Calls', f'=SUMPRODUCT(({col_type}{item_start}:{col_type}{item_end}="Call")*({col_ca}{item_start}:{col_ca}{item_end}="Yes")*({col_amt}{item_start}:{col_amt}{item_end}))'),
        ('Commitment Affecting — Distributions', f'=SUMPRODUCT(({col_type}{item_start}:{col_type}{item_end}="Dist")*({col_ca}{item_start}:{col_ca}{item_end}="Yes")*({col_amt}{item_start}:{col_amt}{item_end}))'),
        ('Commitment Affecting — Subtotal', None),  # =SUM of above 2
        ('Non-Commitment — Calls', f'=SUMPRODUCT(({col_type}{item_start}:{col_type}{item_end}="Call")*({col_ca}{item_start}:{col_ca}{item_end}="No")*({col_amt}{item_start}:{col_amt}{item_end}))'),
        ('Non-Commitment — Distributions', f'=SUMPRODUCT(({col_type}{item_start}:{col_type}{item_end}="Dist")*({col_ca}{item_start}:{col_ca}{item_end}="No")*({col_amt}{item_start}:{col_amt}{item_end}))'),
        ('Non-Commitment — Subtotal', None),
        ('Total Calls', None),
        ('Total Distributions', None),
        ('Net Amount', None),
    ]
    sum_start = r
    for i, (label, formula) in enumerate(summary_items):
        ws.cell(row=r, column=2, value=label).font = _FNT_NORM
        cell = ws.cell(row=r, column=3)
        cell.number_format = _NUM_ACCT
        cell.alignment = _ALIGN_R
        if formula:
            cell.value = formula
        elif i == 2:  # CA Subtotal
            cell.value = f'={col_amt}{sum_start}+{col_amt}{sum_start + 1}'
            cell.font = _FNT_BOLD
            cell.border = Border(top=Side('thin', color=_XS_BDR))
        elif i == 5:  # NC Subtotal
            cell.value = f'={col_amt}{sum_start + 3}+{col_amt}{sum_start + 4}'
            cell.font = _FNT_BOLD
            cell.border = Border(top=Side('thin', color=_XS_BDR))
        elif i == 6:  # Total Calls
            cell.value = f'={col_amt}{sum_start}+{col_amt}{sum_start + 3}'
        elif i == 7:  # Total Dist
            cell.value = f'={col_amt}{sum_start + 1}+{col_amt}{sum_start + 4}'
        elif i == 8:  # Net
            cell.value = f'={col_amt}{r - 2}+{col_amt}{r - 1}'
            cell.font = _FNT_BOLD
            cell.border = _BDR_TOP_DBL
            ws.cell(row=r, column=2).font = _FNT_BOLD
        r += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    # Freeze pane below header info
    ws.freeze_panes = f'A{hdr_row + 1}'


def _xs_wire_sheet(wb, notice):
    """Build Wire Info sheet."""
    ws = wb.create_sheet('Wire Info')
    h = notice['header']
    wires = h.get('wire_info', [])
    wire = wires[0] if wires else None
    net = _pn(h.get('LP_net_amount'))

    ws.merge_cells('A1:B1')
    ws.cell(row=1, column=1, value='Wire Information').font = _FNT_TITLE
    ws.cell(row=1, column=1).fill = _FILL_NAVY
    ws.cell(row=1, column=2).fill = _FILL_NAVY
    ws.row_dimensions[1].height = 30

    r = 3
    ws.cell(row=r, column=1, value='Direction').font = _FNT_BOLD
    ws.cell(row=r, column=2, value='Outbound (Capital Call)' if net and net > 0 else 'Inbound (Distribution)' if net and net < 0 else 'Adjustment')
    r += 1
    ws.cell(row=r, column=1, value='Net Amount').font = _FNT_BOLD
    c = ws.cell(row=r, column=2, value=net)
    c.number_format = _NUM_ACCT

    if not wire:
        r += 2
        ws.cell(row=r, column=1, value='Wire 정보 없음').font = _FNT_SMALL
    else:
        # Beneficiary Bank
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=1, value='Beneficiary Bank (수익자 은행)').font = _FNT_SEC
        ws.cell(row=r, column=1).fill = _FILL_LIGHT
        ws.cell(row=r, column=2).fill = _FILL_LIGHT
        r += 1
        for label, key in [('Bank Name', 'beneficiary_bank_name'), ('Address', 'beneficiary_bank_address'),
                           ('SWIFT / BIC', 'beneficiary_bank_swift_code'), ('ABA Routing', 'beneficiary_bank_aba_routing'),
                           ('Account #', 'beneficiary_bank_account_number')]:
            val = wire.get(key, '') or ''
            if val and val != 'N/A':
                ws.cell(row=r, column=1, value=label).font = _FNT_BOLD
                ws.cell(row=r, column=1).fill = _FILL_GRAY
                ws.cell(row=r, column=2, value=val).font = _FNT_NORM
                if 'account' in key.lower() or 'routing' in key.lower():
                    ws.cell(row=r, column=2).number_format = '@'  # 텍스트 서식 (앞자리 0 보존)
                r += 1

        # Beneficiary
        bn = wire.get('beneficiary_name', '') or wire.get('beneficiary_account_name', '') or ''
        ba = wire.get('beneficiary_account_number', '') or ''
        if (bn and bn != 'N/A') or (ba and ba != 'N/A'):
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=1, value='Beneficiary (수익자)').font = _FNT_SEC
            ws.cell(row=r, column=1).fill = _FILL_LIGHT
            ws.cell(row=r, column=2).fill = _FILL_LIGHT
            r += 1
            if bn and bn != 'N/A':
                ws.cell(row=r, column=1, value='Name').font = _FNT_BOLD
                ws.cell(row=r, column=1).fill = _FILL_GRAY
                ws.cell(row=r, column=2, value=bn)
                r += 1
            if ba and ba != 'N/A':
                ws.cell(row=r, column=1, value='Account #').font = _FNT_BOLD
                ws.cell(row=r, column=1).fill = _FILL_GRAY
                ws.cell(row=r, column=2, value=ba).number_format = '@'
                r += 1

        # Reference / Further Credit
        ref = wire.get('reference', '') or ''
        fc = wire.get('further_credit', '') or ''
        if (ref and ref != 'N/A') or (fc and fc != 'N/A'):
            if ref and ref != 'N/A':
                ws.cell(row=r, column=1, value='Reference').font = _FNT_BOLD
                ws.cell(row=r, column=1).fill = _FILL_GRAY
                ws.cell(row=r, column=2, value=ref)
                r += 1
            if fc and fc != 'N/A':
                ws.cell(row=r, column=1, value='Further Credit').font = _FNT_BOLD
                ws.cell(row=r, column=1).fill = _FILL_GRAY
                ws.cell(row=r, column=2, value=fc)
                r += 1

        # Intermediary
        ib = wire.get('intermediary_bank_name', '') or ''
        if ib and ib != 'N/A':
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=1, value='Intermediary Bank (중개은행)').font = _FNT_SEC
            ws.cell(row=r, column=1).fill = _FILL_LIGHT
            ws.cell(row=r, column=2).fill = _FILL_LIGHT
            r += 1
            for label, key in [('Bank Name', 'intermediary_bank_name'), ('SWIFT / BIC', 'intermediary_swift_code'),
                               ('Account #', 'intermediary_account_number')]:
                val = wire.get(key, '') or ''
                if val and val != 'N/A':
                    ws.cell(row=r, column=1, value=label).font = _FNT_BOLD
                    ws.cell(row=r, column=1).fill = _FILL_GRAY
                    ws.cell(row=r, column=2, value=val)
                    r += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 55


def _xs_commitment_sheet(wb, notices, current_id):
    """Build Commitment sheet with formulas."""
    ws = wb.create_sheet('Commitment')
    if not notices:
        ws.cell(row=1, column=1, value='데이터 없음')
        return

    h0 = notices[0]['header']
    commit = None
    for n in notices:
        c = _pn(n['header'].get('Commitment_original'))
        if c: commit = c

    # Title
    ws.merge_cells('A1:I1')
    ws.cell(row=1, column=1,
        value=f"Commitment History — {h0.get('Underlying_Fund_Name_full', '')}").font = _FNT_TITLE
    ws.cell(row=1, column=1).fill = _FILL_NAVY
    for c in range(2, 10):
        ws.cell(row=1, column=c).fill = _FILL_NAVY
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.cell(row=2, column=1, value=f'Commitment: {commit:,.2f}' if commit else 'Commitment: N/A').font = _FNT_SMALL
    currency = h0.get('Currency', 'USD') or 'USD'
    if currency == 'N/A': currency = 'USD'
    ws.cell(row=2, column=2, value=f'Currency: {currency}').font = _FNT_SMALL
    ws.cell(row=2, column=3, value=f'Notices: {len(notices)}').font = _FNT_SMALL

    # Headers (row 4) — 웹 UI와 동기화된 컬럼명
    headers = ['#', 'Date', 'Type', 'Funded Prior', '약정 소진', 'Funded After', 'Unfunded After', '소진율', 'Δ']
    for ci, txt in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=txt)
    _xs_style_header_row(ws, 4, 9)

    # Data rows
    prev_pct = None
    for i, n in enumerate(notices):
        r = 5 + i
        hh = n['header']
        is_cur = (n['id'] == current_id)
        uf_prior = _pn(hh.get('Unfunded_prior'))
        uf_after = _pn(hh.get('Unfunded_after'))
        funded_prior = (commit - uf_prior) if commit and uf_prior is not None else None
        funded_after = (commit - uf_after) if commit and uf_after is not None else None
        pct = (1 - uf_after / commit) if commit and uf_after is not None else None
        delta = (pct - prev_pct) if pct is not None and prev_pct is not None else (pct if pct is not None and i == 0 else None)

        # 약정 소진 (CA amount)
        ca = 0
        for it in n.get('line_items', []):
            if it.get('is_subtotal'): continue
            if it.get('Commitment_affecting'):
                a = _pn(it.get('LP_signed_amount'))
                if a is not None: ca += a

        ws.cell(row=r, column=1, value=i + 1).alignment = _ALIGN_C
        dt_val = _xs_date(hh.get('issue_date', ''))
        c_date = ws.cell(row=r, column=2, value=dt_val)
        if isinstance(dt_val, dt_date): c_date.number_format = _DATE
        ws.cell(row=r, column=3, value=hh.get('notice_type', '')).alignment = _ALIGN_C
        ws.cell(row=r, column=4, value=funded_prior).number_format = _NUM_ACCT
        ws.cell(row=r, column=4).alignment = _ALIGN_R
        ws.cell(row=r, column=5, value=round(ca, 2) if ca else 0).number_format = _NUM_ACCT
        ws.cell(row=r, column=5).alignment = _ALIGN_R
        ws.cell(row=r, column=6, value=funded_after).number_format = _NUM_ACCT
        ws.cell(row=r, column=6).alignment = _ALIGN_R
        ws.cell(row=r, column=7, value=uf_after).number_format = _NUM_ACCT
        ws.cell(row=r, column=7).alignment = _ALIGN_R
        c_pct = ws.cell(row=r, column=8, value=pct)
        c_pct.number_format = _PCT
        c_pct.alignment = _ALIGN_R
        c_pct.font = _FNT_BOLD
        c_delta = ws.cell(row=r, column=9, value=delta)
        c_delta.number_format = '+0.0%;-0.0%'
        c_delta.alignment = _ALIGN_R

        # Highlight current notice
        if is_cur:
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = _FILL_CUR
                if c in (1, 2, 3): ws.cell(row=r, column=c).font = _FNT_BOLD

        prev_pct = pct

    # Column widths
    for col, w in [('A', 6), ('B', 14), ('C', 12), ('D', 18), ('E', 18), ('F', 18), ('G', 18), ('H', 12), ('I', 10)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'


def _xs_exposure_sheet(wb, notices, current_id, asset_groups=None):
    """Build Exposure sheet with subtotals and formulas."""
    ws = wb.create_sheet('Asset Exposure')
    if not notices:
        ws.cell(row=1, column=1, value='데이터 없음')
        return

    h0 = notices[0]['header']
    n_count = len(notices)

    # Title
    title = f"Asset Exposure Matrix — {h0.get('Underlying_Fund_Name_short', '')} / {h0.get('LP_Name_short', '')}"
    last_col = n_count + 3  # A:Group, B:Raw, C~:notices, last:Total
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=1, column=1, value=title).font = _FNT_TITLE
    ws.cell(row=1, column=1).fill = _FILL_NAVY
    for c in range(2, last_col + 1):
        ws.cell(row=1, column=c).fill = _FILL_NAVY
    ws.row_dimensions[1].height = 30

    # Column headers (row 3)
    ws.cell(row=3, column=1, value='Asset Group')
    ws.cell(row=3, column=2, value='Item Name')
    for ni, n in enumerate(notices):
        dt = n['header'].get('issue_date', '')[:10] or '?'
        tp = n['header'].get('notice_type', '')
        ws.cell(row=3, column=3 + ni, value=f"#{ni + 1} {dt} ({tp})")
    ws.cell(row=3, column=last_col, value='Total')
    _xs_style_header_row(ws, 3, last_col)

    # Build grouped data
    groups = {}  # canonical → {raw → [amts per notice]}
    ag = asset_groups or {}
    ag_reverse = {}
    for canonical, members in ag.items():
        for m in (members or []):
            ag_reverse[m] = canonical

    def _get_canonical(raw):
        if raw in ag_reverse: return ag_reverse[raw]
        return raw  # fallback

    for ni, n in enumerate(notices):
        if n.get('is_voided'): continue
        for it in n.get('line_items', []):
            if it.get('is_subtotal'): continue
            raw = it.get('item_name', '')
            if not raw: continue
            canonical = _get_canonical(raw)
            if canonical not in groups: groups[canonical] = {}
            if raw not in groups[canonical]: groups[canonical][raw] = [0] * n_count
            amt = _pn(it.get('LP_signed_amount')) or 0
            groups[canonical][raw][ni] += amt

    # Write data
    r = 4
    canonicals = sorted(groups.keys(), key=lambda s: s.lower())
    for canonical in canonicals:
        raw_map = groups[canonical]
        raw_names = sorted(raw_map.keys())
        group_start = r

        for raw in raw_names:
            amts = raw_map[raw]
            ws.cell(row=r, column=1, value=canonical if r == group_start else '').font = _FNT_NORM
            ws.cell(row=r, column=2, value=raw).font = _FNT_NORM
            total_col_letter = get_column_letter(last_col)
            first_data_col = get_column_letter(3)
            last_data_col = get_column_letter(2 + n_count)
            for ni, amt in enumerate(amts):
                if amt != 0:
                    c = ws.cell(row=r, column=3 + ni, value=round(amt, 2))
                    c.number_format = _NUM_ACCT
                    c.alignment = _ALIGN_R
            # Total column — formula
            ws.cell(row=r, column=last_col,
                value=f'=SUM({first_data_col}{r}:{last_data_col}{r})').number_format = _NUM_ACCT
            ws.cell(row=r, column=last_col).alignment = _ALIGN_R
            r += 1

        # Subtotal row (only if >1 raw names)
        if len(raw_names) > 1:
            ws.cell(row=r, column=1, value=f'Subtotal: {canonical}').font = _FNT_BOLD
            for ci in range(3, last_col + 1):
                col_l = get_column_letter(ci)
                c = ws.cell(row=r, column=ci,
                    value=f'=SUM({col_l}{group_start}:{col_l}{r - 1})')
                c.number_format = _NUM_ACCT
                c.alignment = _ALIGN_R
                c.font = _FNT_BOLD
                c.border = Border(top=Side('thin', color=_XS_BDR))
            r += 1

    # Grand Total
    r += 1
    ws.cell(row=r, column=1, value='Grand Total').font = Font(name='Calibri', size=10, bold=True, color=_XS_WHITE)
    ws.cell(row=r, column=1).fill = _FILL_NAVY
    for ci in range(2, last_col + 1):
        ws.cell(row=r, column=ci).fill = _FILL_NAVY
        if ci >= 3:
            # Sum all data cells in this column (skip subtotal rows — use raw data)
            col_l = get_column_letter(ci)
            c = ws.cell(row=r, column=ci,
                value=f'=SUMPRODUCT((A4:A{r - 2}<>"")*({col_l}4:{col_l}{r - 2}))')
            c.number_format = _NUM_ACCT
            c.font = Font(name='Calibri', size=10, bold=True, color=_XS_WHITE)
            c.alignment = _ALIGN_R

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35
    for ci in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.freeze_panes = 'C4'


@app.post("/api/export")
async def export_excel(body: dict, request: Request):
    """Generate professional Excel report with openpyxl."""
    user = await get_current_user(request)
    notice_id = body.get("notice_id", "")
    scope = body.get("scope", "current")  # 'current' | 'all'
    fmt = body.get("format", "xlsx")

    # Load current notice
    r = await check_notice_access(notice_id, user)
    if not r:
        raise HTTPException(404, "Notice not found")
    header = r["header"] if isinstance(r["header"], dict) else json.loads(r["header"] or "{}")
    items = r["line_items"] if isinstance(r["line_items"], list) else json.loads(r["line_items"] or "[]")
    _migrate_header_wire(header)
    _normalize_header_wire_beneficiary(header)
    current_notice = {"id": r["id"], "header": header, "line_items": items,
                      "is_voided": bool(r.get("is_voided")), "fileName": r.get("file_name", "")}

    # Collect Fund+LP notices for Exposure/Commitment
    fund_key = header.get("Fund_ID_Key", "") or _make_fund_id_key(header.get("Underlying_Fund_Name_full", ""))
    lp_code = header.get("LP_code", "")
    lp_full = header.get("LP_Name_full", "")

    if is_admin(user):
        all_rows = db_list("notices", order_col="created_at", order_desc=False)
    else:
        all_rows = db_list("notices", order_col="created_at", order_desc=False, org_id=user.get("org_id", ""))

    fund_lp_notices = []
    for nr in all_rows:
        nh = nr["header"] if isinstance(nr["header"], dict) else json.loads(nr["header"] or "{}")
        ni = nr["line_items"] if isinstance(nr["line_items"], list) else json.loads(nr["line_items"] or "[]")
        nfk = nh.get("Fund_ID_Key", "") or _make_fund_id_key(nh.get("Underlying_Fund_Name_full", ""))
        if nfk != fund_key:
            continue
        if lp_code:
            if (nh.get("LP_code", "") or "") != lp_code:
                continue
        else:
            if (nh.get("LP_Name_full", "") or "") != lp_full:
                continue
        _migrate_header_wire(nh)
        _normalize_header_wire_beneficiary(nh)
        fund_lp_notices.append({
            "id": nr["id"], "header": nh, "line_items": ni,
            "is_voided": bool(nr.get("is_voided")),
            "fileName": nr.get("file_name", "")
        })
    # Sort by issue_date
    fund_lp_notices.sort(key=lambda x: x["header"].get("issue_date", "") or "")
    non_voided = [n for n in fund_lp_notices if not n["is_voided"]]

    # Load asset groups
    ag = {}
    try:
        ag_row = db_get("asset_groups", fund_key, id_col="fund_key")
        if ag_row:
            g = ag_row.get("groups")
            if isinstance(g, str): g = json.loads(g)
            if g: ag = g
    except:
        pass

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default Sheet

    if scope == 'current':
        _xs_notice_sheet(wb, current_notice, 'Notice')
    else:
        for i, n in enumerate(non_voided):
            dt = (n["header"].get("issue_date", "") or "")[:10]
            name = f"Notice_{i + 1}_{dt}"[:31]
            _xs_notice_sheet(wb, n, name)

    _xs_wire_sheet(wb, current_notice)
    _xs_exposure_sheet(wb, non_voided, notice_id, ag)
    _xs_commitment_sheet(wb, non_voided, notice_id)

    # Sheet tab colors
    for ws in wb.worksheets:
        if 'Notice' in ws.title:
            ws.sheet_properties.tabColor = _XS_NAVY[2:]
        elif ws.title == 'Wire Info':
            ws.sheet_properties.tabColor = _XS_GREEN[2:]
        elif ws.title == 'Asset Exposure':
            ws.sheet_properties.tabColor = _XS_AMBER[2:]
        elif ws.title == 'Commitment':
            ws.sheet_properties.tabColor = _XS_NAVY[2:]

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Filename
    fund_short = (header.get("Underlying_Fund_Name_short", "") or "Fund").replace("/", "_")[:30]
    lp_short = (lp_code or header.get("LP_Name_short", "") or "LP").replace("/", "_")[:30]
    dt_str = (header.get("issue_date", "") or "export")[:10]
    filename = f"{fund_short}_{lp_short}_{dt_str if scope == 'current' else 'ALL'}.xlsx"

    from starlette.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", "8000"))
    print(f"\n🚀 LP Notice Analyzer Backend")
    print(f"   API: http://localhost:{port}/api/health")
    print(f"   App: http://localhost:{port}/app\n")
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
